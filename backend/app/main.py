"""
Djeliya — Serveur de transcription et d'analyse pour la recherche qualitative.
Déployé sur Railway. Whisper (FR/EN + langues locales expérimentales),
diarisation des locuteurs, analyse qualitative méthode Gioia, comptes et
corpus partagés en équipe, fiabilité inter-codeurs.
"""

import os
import re
import json
import uuid
import secrets
import hashlib
import tempfile
import threading
import requests
from datetime import datetime, timedelta, timezone
from itertools import combinations

from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel

from .db import (
    init_db, get_session, Utilisateur, Corpus, MembreCorpus, Entretien, Codage,
    ContributionLangue, Forfait, MouvementCredit, GuideEntretien, Commande,
    EtudeQuantitative, AnalyseQuantitative, ADMIN_EMAIL,
)
from .auth import hacher_mot_de_passe, verifier_mot_de_passe, creer_jeton, utilisateur_courant
from .email_utils import envoyer_code_verification, envoyer_code_reinitialisation, EMAIL_CONFIGURE
from .cgu import CGU_TEXTE, CGU_VERSION
from .exports import (
    generer_docx_entretien, generer_xlsx_entretien, generer_docx_etude, generer_xlsx_etude, generer_docx_guide,
    generer_docx_etude_quant, generer_xlsx_template_questionnaire, generer_docx_analyse_quant, generer_xlsx_analyse_quant,
)
from .stats_quant import analyser_donnees

# ----------------------------------------------------------------- config
WHISPER_MODEL = os.getenv("WHISPER_MODEL", "small")
COMPUTE_TYPE = os.getenv("COMPUTE_TYPE", "int8")
MAX_UPLOAD_MB = int(os.getenv("MAX_UPLOAD_MB", "200"))
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "*").split(",")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
ANALYSE_MODEL = os.getenv("ANALYSE_MODEL", "claude-sonnet-5")
HF_TOKEN = os.getenv("HF_TOKEN")
CREDITS_ESSAI_GRATUIT = float(os.getenv("CREDITS_ESSAI_GRATUIT", "5"))
COUT_CREDIT_TRANSCRIPTION = float(os.getenv("COUT_CREDIT_TRANSCRIPTION", "1"))
COUT_CREDIT_ANALYSE = float(os.getenv("COUT_CREDIT_ANALYSE", "2"))
COUT_CREDIT_GUIDE = float(os.getenv("COUT_CREDIT_GUIDE", "1"))
COUT_CREDIT_ETUDE_QUANT = float(os.getenv("COUT_CREDIT_ETUDE_QUANT", "3"))
COUT_CREDIT_ANALYSE_QUANT = float(os.getenv("COUT_CREDIT_ANALYSE_QUANT", "2"))

# Recharge de crédits à la carte (paiement réel par les utilisateurs)
PRIX_CREDIT_FCFA = int(os.getenv("PRIX_CREDIT_FCFA", "150"))
CREDITS_MIN_RECHARGE = float(os.getenv("CREDITS_MIN_RECHARGE", "10"))
PAYMENT_PROVIDER = os.getenv("PAYMENT_PROVIDER")  # ex. "cinetpay", "paydunya" — à confirmer
BACKEND_PUBLIC_URL = os.getenv("BACKEND_PUBLIC_URL", "").rstrip("/")

# PayDunya exige TROIS identifiants pour authentifier chaque requête — la clé
# maîtresse seule ne suffit jamais (voir developers.paydunya.com).
PAYDUNYA_MASTER_KEY = os.getenv("PAYDUNYA_MASTER_KEY")
PAYDUNYA_PRIVATE_KEY = os.getenv("PAYDUNYA_PRIVATE_KEY")
PAYDUNYA_TOKEN = os.getenv("PAYDUNYA_TOKEN")
PAYDUNYA_MODE = os.getenv("PAYDUNYA_MODE", "test")  # "test" (bac à sable) ou "live" — jamais "live" par défaut


def _paydunya_pret() -> bool:
    return bool(PAYDUNYA_MASTER_KEY and PAYDUNYA_PRIVATE_KEY and PAYDUNYA_TOKEN and BACKEND_PUBLIC_URL)


def _paydunya_base_url() -> str:
    return "https://app.paydunya.com/api/v1" if PAYDUNYA_MODE == "live" else "https://app.paydunya.com/sandbox-api/v1"


def _paydunya_headers() -> dict:
    return {
        "Content-Type": "application/json",
        "PAYDUNYA-MASTER-KEY": PAYDUNYA_MASTER_KEY,
        "PAYDUNYA-PRIVATE-KEY": PAYDUNYA_PRIVATE_KEY,
        "PAYDUNYA-TOKEN": PAYDUNYA_TOKEN,
    }

app = FastAPI(title="Djeliya API", version="0.3.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)


@app.on_event("startup")
def _demarrage():
    init_db()


LANGUES_SUPPORTEES = {
    "auto": "Détection automatique",
    "fr": "Français",
    "en": "Anglais",
    "dyu": "Dioula (expérimental — nécessite un modèle affiné)",
    "bci": "Baoulé (expérimental — nécessite un modèle affiné)",
}

# ----------------------------------------------------------------- modèles paresseux
_whisper_model = None
_whisper_lock = threading.Lock()


def get_whisper():
    global _whisper_model
    with _whisper_lock:
        if _whisper_model is None:
            from faster_whisper import WhisperModel
            _whisper_model = WhisperModel(WHISPER_MODEL, compute_type=COMPUTE_TYPE)
        return _whisper_model


_anthropic_client = None


def get_anthropic():
    global _anthropic_client
    if not ANTHROPIC_API_KEY:
        raise RuntimeError(
            "ANTHROPIC_API_KEY n'est pas configurée — ajoute-la dans les variables Railway "
            "pour activer l'analyse qualitative."
        )
    if _anthropic_client is None:
        from anthropic import Anthropic
        _anthropic_client = Anthropic(api_key=ANTHROPIC_API_KEY)
    return _anthropic_client


_diarisation_pipeline = None
_diarisation_lock = threading.Lock()


def get_diarisation():
    global _diarisation_pipeline
    if not HF_TOKEN:
        raise RuntimeError(
            "HF_TOKEN n'est pas configuré — la diarisation des locuteurs est désactivée. "
            "Accepte les conditions du modèle pyannote/speaker-diarization-3.1 sur Hugging Face "
            "puis ajoute ton jeton dans les variables Railway pour l'activer."
        )
    with _diarisation_lock:
        if _diarisation_pipeline is None:
            from pyannote.audio import Pipeline
            _diarisation_pipeline = Pipeline.from_pretrained(
                "pyannote/speaker-diarization-3.1", use_auth_token=HF_TOKEN
            )
        return _diarisation_pipeline


# ----------------------------------------------------------------- utilitaires
def _nom_fichier(texte: str) -> str:
    """Translittère en ASCII pur : les en-têtes HTTP n'acceptent pas les accents,
    et un nom de fichier avec un « é » brut cassait le téléchargement."""
    import unicodedata
    ascii_txt = unicodedata.normalize("NFKD", texte or "").encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^\w\-]+", "_", ascii_txt).strip("_") or "export"


def _entretien_vers_dict(e: Entretien) -> dict:
    segments = list(e.segments or [])
    if e.locuteurs:
        for s in segments:
            meilleur, recouvrement_max = None, 0.0
            for loc in e.locuteurs:
                debut = max(s.get("debut", 0), loc["debut"])
                fin = min(s.get("fin", s.get("debut", 0)), loc["fin"])
                recouvrement = max(0.0, fin - debut)
                if recouvrement > recouvrement_max:
                    meilleur, recouvrement_max = loc["locuteur"], recouvrement
            if meilleur:
                s["locuteur"] = meilleur

    return {
        "id": e.id,
        "titre": e.titre,
        "langue": e.langue,
        "statut": e.statut,
        "erreur": e.erreur,
        "note": e.note,
        "langue_detectee": e.langue_detectee,
        "segments": segments,
        "diarisation_disponible": bool(e.locuteurs),
        "corpus_id": e.corpus_id,
        "analyse_statut": e.analyse_statut,
        "analyse": e.analyse,
        "analyse_erreur": e.analyse_erreur,
        "analyse_contexte": e.analyse_contexte,
        "analyse_methode": e.analyse_methode,
        "analyse_langue": e.analyse_langue,
        "analyse_modele": e.analyse_modele,
        "cree_le": e.cree_le.isoformat() if e.cree_le else None,
    }


def _verifier_acces(session, entretien: Entretien, user: Utilisateur):
    if entretien.proprietaire_id == user.id:
        return
    if entretien.corpus_id:
        membre = session.query(MembreCorpus).filter_by(
            corpus_id=entretien.corpus_id, utilisateur_id=user.id
        ).first()
        if membre:
            return
    raise HTTPException(403, "Tu n'as pas accès à cet entretien.")


def _generer_code_invitation() -> str:
    return secrets.token_hex(4).upper()


# ----------------------------------------------------------------- schémas
class InscriptionIn(BaseModel):
    email: str
    mot_de_passe: str
    nom: str = ""
    accepte_cgu: bool = False


class ModifierCompteIn(BaseModel):
    nom: str | None = None
    mot_de_passe_actuel: str | None = None
    nouveau_mot_de_passe: str | None = None


class SupprimerCompteIn(BaseModel):
    mot_de_passe: str


class ConnexionIn(BaseModel):
    email: str
    mot_de_passe: str


class VerifierEmailIn(BaseModel):
    email: str
    code: str


class RenvoyerCodeIn(BaseModel):
    email: str


class MotDePasseOublieIn(BaseModel):
    email: str


class ReinitialiserMdpIn(BaseModel):
    email: str
    code: str
    nouveau_mot_de_passe: str


class CreditsIn(BaseModel):
    delta: float
    motif: str = ""


class DefinirCreditsIn(BaseModel):
    valeur: float
    motif: str = ""


class GuideEntretienIn(BaseModel):
    theme: str
    question_recherche: str = ""
    langue: str = "fr"


class RechargeIn(BaseModel):
    credits: float


class ForfaitIn(BaseModel):
    nom: str
    prix_fcfa: int
    credits_inclus: float
    description: str = ""


class AttribuerForfaitIn(BaseModel):
    forfait_id: str


class CorpusIn(BaseModel):
    nom: str


class RejoindreIn(BaseModel):
    code: str


class AnalyseCorpusIn(BaseModel):
    contexte: str = ""
    methode: str = "gioia"
    langue: str = "fr"


class CodageIn(BaseModel):
    segment_index: int
    code: str


class PreferencesIn(BaseModel):
    contribution_langues_locales: bool


class ContributionIn(BaseModel):
    langue: str
    texte_original: str
    texte_corrige: str


class SegmentsIn(BaseModel):
    segments: list


# ----------------------------------------------------------------- routes publiques
@app.get("/", response_class=HTMLResponse)
def accueil():
    """Page d'accueil sommaire du serveur — sert aussi de page de retour après un
    paiement PayDunya (return_url), pour éviter d'afficher un 404 brut au chercheur."""
    return """<!DOCTYPE html>
<html lang="fr"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Djeliya</title>
<style>
  body { background:#0E1226; color:#F4EFE3; font-family:-apple-system,Segoe UI,Roboto,sans-serif;
         display:flex; align-items:center; justify-content:center; min-height:100vh; margin:0; text-align:center; padding:24px; }
  .card { max-width:380px; }
  h1 { color:#E4B04A; font-size:22px; margin-bottom:8px; }
  p { color:#B7B2A6; font-size:15px; line-height:1.5; }
</style></head>
<body><div class="card">
  <h1>Djeliya</h1>
  <p>Le serveur fonctionne normalement.</p>
  <p>Si tu reviens d'un paiement, retourne simplement dans l'application Djeliya sur ton téléphone —
  ton solde de crédits se met à jour automatiquement une fois le paiement confirmé.</p>
</div></body></html>"""


@app.get("/health")
def health():
    return {
        "status": "ok",
        "model": WHISPER_MODEL,
        "analyse_disponible": bool(ANTHROPIC_API_KEY),
        "diarisation_disponible": bool(HF_TOKEN),
        "email_configure": EMAIL_CONFIGURE,
        "time": datetime.now(timezone.utc).isoformat(),
        # Fournis automatiquement par Railway lors du build — permet de vérifier
        # sans ambiguïté que le déploiement en ligne correspond bien au dernier
        # commit poussé sur GitHub (comparer avec le SHA affiché sur github.com).
        "commit_deploye": os.getenv("RAILWAY_GIT_COMMIT_SHA", "inconnu — pas sur Railway ou variable absente"),
    }


@app.get("/api/languages")
def languages():
    return LANGUES_SUPPORTEES


@app.get("/api/methodes")
def methodes():
    return {k: {"label": v["label"], "reference": v["reference"]} for k, v in METHODES_ANALYSE.items()}


@app.get("/api/cgu")
def cgu():
    return {"version": CGU_VERSION, "texte": CGU_TEXTE}


# ----------------------------------------------------------------- guide d'entretien
def _guide_vers_dict(g: GuideEntretien) -> dict:
    return {
        "id": g.id, "theme": g.theme, "question_recherche": g.question_recherche,
        "langue": g.langue, "statut": g.statut, "guide": g.guide, "erreur": g.erreur,
        "modele": g.modele, "cree_le": g.cree_le.isoformat() if g.cree_le else None,
    }


@app.get("/api/guides")
def lister_guides(user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        guides = session.query(GuideEntretien).filter_by(proprietaire_id=user.id).order_by(GuideEntretien.cree_le.desc()).all()
        return [_guide_vers_dict(g) for g in guides]
    finally:
        session.close()


@app.get("/api/guides/{guide_id}")
def detail_guide(guide_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        g = session.get(GuideEntretien, guide_id)
        if not g or g.proprietaire_id != user.id:
            raise HTTPException(404, "Guide introuvable.")
        return _guide_vers_dict(g)
    finally:
        session.close()


@app.post("/api/guides")
def creer_guide(payload: GuideEntretienIn, user: Utilisateur = Depends(utilisateur_courant)):
    if not payload.theme.strip():
        raise HTTPException(422, "Le thème de recherche est requis.")
    theme = _nettoyer_texte_utilisateur(payload.theme.strip())
    question_recherche = _nettoyer_texte_utilisateur(payload.question_recherche.strip())
    langue = payload.langue if payload.langue in ("fr", "en") else "fr"
    session = get_session()
    try:
        if not user.email_verifie:
            raise HTTPException(403, "Vérifie ton adresse e-mail avant de générer un guide d'entretien.")
        if not user.est_admin and user.credits < COUT_CREDIT_GUIDE:
            raise HTTPException(402, "Crédits insuffisants. Consulte les forfaits disponibles dans l'app.")
        if not user.est_admin:
            u = session.get(Utilisateur, user.id)
            u.credits -= COUT_CREDIT_GUIDE
            session.add(MouvementCredit(
                id=uuid.uuid4().hex[:16], utilisateur_id=user.id, delta=-COUT_CREDIT_GUIDE,
                motif=f"Guide d'entretien — {theme[:60]}",
            ))
        guide_id = uuid.uuid4().hex[:16]
        g = GuideEntretien(
            id=guide_id, proprietaire_id=user.id, theme=theme,
            question_recherche=question_recherche, langue=langue, statut="en_cours",
        )
        session.add(g)
        session.commit()
    finally:
        session.close()

    threading.Thread(target=_run_guide, args=(guide_id, theme, question_recherche, langue, user.id), daemon=True).start()
    return {"id": guide_id, "statut": "en_cours"}


@app.delete("/api/guides/{guide_id}")
def supprimer_guide(guide_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        g = session.get(GuideEntretien, guide_id)
        if not g or g.proprietaire_id != user.id:
            raise HTTPException(404, "Guide introuvable.")
        session.delete(g)
        session.commit()
        return {"ok": True}
    finally:
        session.close()


@app.get("/api/guides/{guide_id}/export/docx")
def export_docx_guide(guide_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        g = session.get(GuideEntretien, guide_id)
        if not g or g.proprietaire_id != user.id:
            raise HTTPException(404, "Guide introuvable.")
        if g.statut != "termine" or not g.guide:
            raise HTTPException(409, "Ce guide n'est pas encore prêt.")
        data = _guide_vers_dict(g)
    finally:
        session.close()
    buf = generer_docx_guide(data)
    nom = _nom_fichier(data["theme"])
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="guide_{nom}.docx"'},
    )


SCHEMA_GUIDE = """{
  "titre": "titre concis et professionnel du guide d'entretien",
  "informations_pratiques": {
    "type_entretien": "ex. entretien semi-directif individuel",
    "duree_estimee": "ex. 45 à 60 minutes",
    "population_cible": "description du profil des personnes à interroger",
    "materiel_recommande": "ex. enregistreur, formulaire de consentement, carnet de notes"
  },
  "preambule": "texte complet à lire au participant en début d'entretien : présentation du chercheur et de l'étude, objectif, durée, caractère volontaire, confidentialité et anonymisation, demande d'autorisation d'enregistrement, droit de ne pas répondre ou d'arrêter à tout moment",
  "sections": [
    {
      "titre": "libellé de la section (ex. Mise en confiance, ou le nom d'un thème précis)",
      "objectif": "ce que cette section vise à recueillir, en 1 phrase",
      "questions": [
        {"question": "question ouverte posée telle quelle", "relances": ["relance 1 si la réponse est courte", "relance 2 pour approfondir"]}
      ]
    }
  ],
  "conseils_methodologiques": "conseils concrets pour la conduite de l'entretien : écoute active, gestion des silences, neutralité du chercheur, prise de notes, reformulation",
  "note_methodologique": "paragraphe de niveau doctoral situant le guide dans son cadre théorique (type d'entretien, logique de l'entonnoir du général au particulier, typologie des questions mobilisée) avec au moins une référence bibliographique reconnue (ex. Kvale & Brinkmann, 2009 ; Patton, 2002 ; Blanchet & Gotman, 2007)",
  "grille_coherence": [
    {"question": "reprend EXACTEMENT le texte d'une question du guide ci-dessus", "dimension_visee": "la dimension ou l'axe théorique du sujet que cette question vise à explorer", "justification": "en une phrase, pourquoi cette question permet d'explorer cette dimension précise"}
  ]
}"""


def _construire_prompt_guide(theme: str, question_recherche: str, langue: str) -> str:
    consigne_langue = "Rédige l'intégralité du guide en anglais." if langue == "en" else "Rédige l'intégralité du guide en français."
    return f"""Tu es méthodologue qualitatif spécialisé en sciences de gestion et organisation, de niveau \
recherche doctorale, expert en conception de guides d'entretien semi-directifs.

Conçois un guide d'entretien professionnel, complet et directement utilisable sur le terrain, à partir \
des éléments suivants fournis par le chercheur :

Thème de recherche : {theme}
Question de recherche : {question_recherche or "non précisée — déduis un axe pertinent à partir du thème"}

Le guide doit suivre la structure classique en entonnoir (questions générales puis progressivement plus \
précises) : une phase de mise en confiance, puis 3 à 6 sections thématiques couvrant les dimensions \
pertinentes du sujet, puis une clôture. Chaque question ouverte doit être accompagnée de relances \
possibles. Le guide doit être directement utilisable par un chercheur sur le terrain, sans jargon inutile \
dans les questions elles-mêmes (le jargon méthodologique reste réservé à la note méthodologique).

Produis également une grille de cohérence ("grille_coherence") qui reprend CHAQUE question principale du \
guide (pas les relances) et l'associe explicitement à la dimension théorique du sujet qu'elle vise à \
explorer, avec une justification courte. Cette grille doit permettre à un chercheur ou un directeur de \
recherche de vérifier d'un coup d'œil qu'aucune dimension du thème n'est oubliée et qu'aucune question \
n'est redondante avec une autre.

{consigne_langue}

{CONSIGNE_ORIGINALITE}

Point d'attention supplémentaire pour ce guide : plusieurs étudiants peuvent soumettre des thèmes très \
proches (ex. plusieurs mémoires sur le même sujet de recherche dans une même promotion). Formule donc \
les questions et les intitulés de section de façon distinctive, en t'appuyant précisément sur les mots \
et l'angle propres à CE thème et CETTE question de recherche fournis, plutôt que sur une trame \
générique de guide d'entretien qui reviendrait à l'identique pour n'importe quel autre thème voisin.

Réponds UNIQUEMENT avec un objet JSON strictement conforme à ce schéma — aucun texte avant ou après, \
aucune balise markdown. Assure-toi que le JSON est syntaxiquement valide : échappe tous les guillemets \
internes aux chaînes de caractères (\\") et ne laisse aucune chaîne non terminée :
{SCHEMA_GUIDE}"""


def _valider_guide(guide: dict):
    for cle in ("titre", "preambule", "sections", "conseils_methodologiques", "note_methodologique", "grille_coherence"):
        if cle not in guide:
            raise ValueError(f"Réponse du modèle incomplète (« {cle} » manquant).")
    if not guide["sections"]:
        raise ValueError("Le guide généré ne contient aucune section.")


def _run_guide(guide_id: str, theme: str, question_recherche: str, langue: str, payeur_id: str):
    session = get_session()
    try:
        client = get_anthropic()
        prompt = _construire_prompt_guide(theme, question_recherche, langue)
        resp = client.messages.create(
            model=ANALYSE_MODEL,
            max_tokens=9000,
            messages=[{"role": "user", "content": prompt}],
        )
        texte = "".join(b.text for b in resp.content if getattr(b, "type", None) == "text")
        guide = _extraire_json(texte)
        _valider_guide(guide)

        g = session.get(GuideEntretien, guide_id)
        g.guide = guide
        g.statut = "termine"
        g.modele = ANALYSE_MODEL
        session.commit()
    except Exception as ex:  # noqa: BLE001
        g = session.get(GuideEntretien, guide_id)
        if g:
            g.statut = "erreur"
            g.erreur = str(ex)
            u = session.get(Utilisateur, payeur_id)
            if u and not u.est_admin:
                u.credits += COUT_CREDIT_GUIDE
                session.add(MouvementCredit(
                    id=uuid.uuid4().hex[:16], utilisateur_id=u.id, delta=COUT_CREDIT_GUIDE,
                    motif="Remboursement — échec de la génération du guide",
                ))
            session.commit()
    finally:
        session.close()


# ----------------------------------------------------------------- étude quantitative
# Références méthodologiques réellement vérifiées (jamais générées par le modèle) —
# seules citations exactes garanties authentiques dans le rapport, pour éviter tout
# risque de référence fabriquée sur la partie théorique, spécifique à chaque thème.
REFERENCES_APA_METHODO_QUANT = [
    "Cronbach, L. J. (1951). Coefficient alpha and the internal structure of tests. Psychometrika, 16(3), 297-334.",
    "Nunnally, J. C., & Bernstein, I. H. (1994). Psychometric theory (3rd ed.). McGraw-Hill.",
    "Fornell, C., & Larcker, D. F. (1981). Evaluating structural equation models with unobservable variables and measurement error. Journal of Marketing Research, 18(1), 39-50.",
    "Churchill, G. A. (1979). A paradigm for developing better measures of marketing constructs. Journal of Marketing Research, 16(1), 64-73.",
    "Hair, J. F., Black, W. C., Babin, B. J., & Anderson, R. E. (2019). Multivariate data analysis (8th ed.). Cengage.",
]

SCHEMA_ETUDE_ETAPE1A = """{
  "titre": "titre concis et professionnel de l'étude (15 mots maximum)",
  "cadre_theorique": "développement de niveau doctoral, 350 à 500 mots STRICTEMENT (soit l'ordre de grandeur conventionnel d'une section de cadre théorique dans un article scientifique), situant le sujet dans un ou plusieurs cadres théoriques reconnus du champ disciplinaire concerné",
  "concepts_theoriques": [
    {"concept": "nom de la théorie ou du courant mobilisé (ex. Théorie de l'échange social)", "auteur_associe": "nom du ou des auteurs fondateurs largement reconnus de cette théorie (ex. Blau, 1964)"}
  ]
}"""

SCHEMA_ETUDE_ETAPE1B = """{
  "revue_litterature": "synthèse de niveau doctoral, 350 à 500 mots STRICTEMENT (ordre de grandeur conventionnel d'une revue de littérature dans un article scientifique), structurée par grands axes/débats"
}"""

SCHEMA_ETUDE_ETAPE2 = """{
  "methodologie": {
    "type_etude": "ex. étude quantitative transversale par questionnaire auto-administré (1 phrase)",
    "population_cible": "description précise du profil des répondants visés (1 à 2 phrases)",
    "echantillon": "taille d'échantillon recommandée avec justification et méthode d'échantillonnage proposée (2 à 3 phrases maximum)",
    "hypotheses": [{"code": "H1", "enonce": "hypothèse testable formulée précisément et concisément, reliant les variables (1 phrase par hypothèse, 5 hypothèses maximum)"}],
    "variables": [{"nom": "nom du construit/variable (6 variables maximum au total)", "type": "indépendante|dépendante|médiatrice|modératrice|de contrôle", "definition": "définition opérationnelle en 1 phrase courte"}]
  }
}"""

SCHEMA_ETUDE_ETAPE3 = """{
  "questionnaire": {
    "sections": [
      {
        "titre": "libellé de section",
        "variable_associee": "nom EXACT d'une variable listée dans la méthodologie si cette section mesure ce construit par échelle de Likert, sinon omettre ce champ",
        "items": [
          {"code": "Q1", "libelle": "énoncé exact de la question ou de l'item, une phrase concise", "type": "choix_unique|choix_multiple|echelle_likert|numerique|texte_libre", "options": ["modalités si choix_unique/choix_multiple, ou les 5 libellés si echelle_likert"], "echelle_min": 1, "echelle_max": 5}
        ]
      }
    ]
  }
}"""
# Limite volontaire : 6 sections maximum, 5 items maximum par section (indiqué au
# modèle dans les instructions de l'étape, pas dans le schéma lui-même).

SCHEMA_ETUDE_ETAPE4 = """{
  "note_methodologique": "paragraphe de niveau doctoral justifiant les choix méthodologiques (type d'échelle, structure du questionnaire, validité de construit)"
}"""


def _nettoyer_texte_utilisateur(texte: str) -> str:
    """Neutralise les guillemets doubles (et variantes typographiques) dans un
    texte saisi par l'utilisateur avant de l'injecter dans un prompt IA — un
    guillemet repris tel quel par le modèle dans sa réponse JSON (ex. dans un
    titre reprenant le thème) peut casser le parsing strict si mal échappé."""
    if not texte:
        return texte
    for car in ('"', "\u201c", "\u201d", "\u201e", "\u00ab", "\u00bb"):
        texte = texte.replace(car, "'")
    return texte


def _prompt_etude_base(theme: str, question_recherche: str, langue: str) -> str:
    consigne_langue = "Rédige l'intégralité en anglais." if langue == "en" else "Rédige l'intégralité en français."
    return f"""Tu es méthodologue quantitatif et directeur de recherche, de niveau recherche doctorale, \
expert en sciences de gestion, économie, sociologie ou sciences connexes selon le thème fourni.

Thème de recherche : {theme}
Question de recherche : {question_recherche or "non précisée — déduis un axe pertinent à partir du thème"}

{consigne_langue}

{CONSIGNE_ORIGINALITE}"""


def _valider_etape(contenu: dict, cles: tuple):
    for cle in cles:
        if cle not in contenu:
            raise ValueError(f"Réponse du modèle incomplète (« {cle} » manquant).")


def _appel_ia_json(prompt: str, max_tokens: int) -> dict:
    client = get_anthropic()
    resp = client.messages.create(model=ANALYSE_MODEL, max_tokens=max_tokens, messages=[{"role": "user", "content": prompt}])
    texte = "".join(b.text for b in resp.content if getattr(b, "type", None) == "text")
    if not texte.strip():
        # Une réponse vide (aucun bloc texte) plutôt qu'un JSON tronqué ou mal formé —
        # cas distinct qu'on isole pour un diagnostic clair au lieu d'une erreur JSON
        # cryptique ("Expecting value: line 1 column 1"), et qui bénéficie de la même
        # reprise automatique que les autres échecs de génération.
        raise ValueError(f"Réponse vide du modèle (motif d'arrêt : {getattr(resp, 'stop_reason', 'inconnu')}).")
    return _extraire_json(texte)


def _appel_etape_avec_reprise(prompt: str, max_tokens: int, verifier) -> dict:
    """Exécute une étape de génération avec une nouvelle tentative automatique si
    la réponse échoue à la validation (« verifier » lève une exception sinon) —
    absorbe les échecs ponctuels de génération sans faire perdre de crédit à
    l'utilisateur pour un simple aléa d'un seul appel au modèle."""
    try:
        resultat = _appel_ia_json(prompt, max_tokens)
        verifier(resultat)
        return resultat
    except Exception:  # noqa: BLE001
        prompt_renforce = f"{prompt}\n\nATTENTION : ta précédente tentative était incomplète ou mal formée. Respecte scrupuleusement le schéma demandé, sans en omettre aucune partie."
        resultat = _appel_ia_json(prompt_renforce, max_tokens)
        verifier(resultat)
        return resultat


def _run_etude_quant(etude_id: str, theme: str, question_recherche: str, langue: str, payeur_id: str):
    """Génère l'étude en 4 étapes indépendantes (cadre théorique, méthodologie,
    questionnaire, note méthodologique), chacune avec sa propre marge de tokens —
    un thème riche produit trop de contenu pour tenir en un seul appel sans risquer
    une troncature en plein milieu du JSON."""
    session = get_session()
    contenu: dict = {}
    try:
        base = _prompt_etude_base(theme, question_recherche, langue)

        e = session.get(EtudeQuantitative, etude_id)
        consigne_integrite = (
            "Consigne d'intégrité scientifique impérative : ne fabrique JAMAIS de référence "
            "bibliographique précise (auteur + année + titre exact d'article) qui n'existerait pas "
            "réellement. Reste au niveau des courants théoriques et notions établies du champ dans le "
            "texte ; ne nomme dans « concepts_theoriques » que des théories réellement célèbres et "
            "incontestables du champ, avec leur(s) auteur(s) fondateur(s) largement reconnu(s).\n\n"
            "Consigne de citation dans le texte : quand tu mobilises une théorie réellement célèbre et "
            "incontestable (ex. théorie de l'échange social, modèle de l'engagement organisationnel), "
            "cite-la au format APA standard « Auteur (Année) » directement dans le corps du texte — pas "
            "seulement le nom de l'auteur seul — mais UNIQUEMENT si tu es certain que cette année de "
            "publication est exacte et largement citée dans le champ (ex. Blau (1964), Meyer et Allen "
            "(1991)) ; en cas de doute sur l'année exacte, nomme la théorie et l'auteur sans donner "
            "d'année plutôt que de risquer une date inventée."
        )

        e.etape = "cadre"
        session.commit()
        prompt1a = f"{base}\n\n{consigne_integrite}\n\nRéponds UNIQUEMENT en JSON conforme à ce schéma :\n{SCHEMA_ETUDE_ETAPE1A}"
        etape1a = _appel_etape_avec_reprise(prompt1a, 3000, lambda r: _valider_etape(r, ("titre", "cadre_theorique")))
        contenu.update(etape1a)
        e.contenu = dict(contenu)
        session.commit()

        e.etape = "revue"
        session.commit()
        contexte1b = f"{base}\n\nCadre théorique déjà établi :\n{contenu['cadre_theorique'][:1500]}"
        prompt1b = f"{contexte1b}\n\n{consigne_integrite}\n\nRéponds UNIQUEMENT en JSON conforme à ce schéma :\n{SCHEMA_ETUDE_ETAPE1B}"
        etape1b = _appel_etape_avec_reprise(prompt1b, 3000, lambda r: _valider_etape(r, ("revue_litterature",)))
        contenu.update(etape1b)
        e.contenu = dict(contenu)
        session.commit()

        e.etape = "methodologie"
        session.commit()
        contexte2 = f"{base}\n\nCadre théorique déjà établi :\n{contenu['cadre_theorique'][:1500]}"
        instructions2 = (
            "Conçois la méthodologie quantitative de cette étude, cohérente avec le cadre théorique "
            "ci-dessus. Chaque hypothèse doit relier des variables précisément définies."
        )
        prompt2 = f"{contexte2}\n\n{instructions2}\n\nRéponds UNIQUEMENT en JSON conforme à ce schéma :\n{SCHEMA_ETUDE_ETAPE2}"
        etape2 = _appel_etape_avec_reprise(prompt2, 4500, lambda r: _valider_etape(r, ("methodologie",)))
        contenu.update(etape2)
        e.contenu = dict(contenu)
        session.commit()

        e.etape = "questionnaire"
        session.commit()
        contexte3 = f"{base}\n\nMéthodologie déjà établie (hypothèses et variables) :\n{json.dumps(contenu['methodologie'], ensure_ascii=False)}"
        instructions3 = (
            "Conçois le questionnaire complet mesurant ces variables : items clairs, échelles de "
            "Likert à 5 points cohérentes (3 à 5 items par construit mesuré par échelle, pour "
            "permettre un calcul ultérieur de fiabilité), plus les variables de contrôle/"
            "sociodémographiques pertinentes. Limite stricte : 6 sections maximum, 5 items maximum "
            "par section — reste concis, un questionnaire trop long décourage les répondants."
        )

        def _verifier_questionnaire(r):
            _valider_etape(r, ("questionnaire",))
            if not r["questionnaire"].get("sections"):
                raise ValueError("Le questionnaire généré ne contient aucune section.")

        prompt3 = f"{contexte3}\n\n{instructions3}\n\nRéponds UNIQUEMENT en JSON conforme à ce schéma :\n{SCHEMA_ETUDE_ETAPE3}"
        etape3 = _appel_etape_avec_reprise(prompt3, 7000, _verifier_questionnaire)
        contenu.update(etape3)
        e.contenu = dict(contenu)
        session.commit()

        e.etape = "references"
        session.commit()
        instructions4 = (
            "Rédige la note méthodologique justifiant les choix méthodologiques (type d'échelle, "
            "structure du questionnaire, validité de construit), sans citer de référence précise "
            "(les références méthodologiques sont ajoutées séparément par l'application)."
        )
        prompt4 = f"{base}\n\n{instructions4}\n\nRéponds UNIQUEMENT en JSON conforme à ce schéma :\n{SCHEMA_ETUDE_ETAPE4}"
        etape4 = _appel_etape_avec_reprise(prompt4, 2500, lambda r: _valider_etape(r, ("note_methodologique",)))
        contenu.update(etape4)

        # Table de références APA : uniquement des sources réellement vérifiées —
        # le socle méthodologique fixe, plus les concepts théoriques nommés par le
        # modèle à l'étape 1, présentés comme à vérifier par le chercheur (jamais
        # comme des citations exactes garanties).
        contenu["references_apa"] = {
            "methodologie": list(REFERENCES_APA_METHODO_QUANT),
            "concepts_a_referencer": contenu.get("concepts_theoriques", []),
        }

        e.contenu = contenu
        e.statut = "termine"
        e.etape = "termine"
        e.modele = ANALYSE_MODEL
        session.commit()
    except Exception as ex:  # noqa: BLE001
        e = session.get(EtudeQuantitative, etude_id)
        if e:
            e.statut = "erreur"
            e.erreur = str(ex)
            e.contenu = contenu or e.contenu  # conserve ce qui a déjà été généré avec succès
            u = session.get(Utilisateur, payeur_id)
            if u and not u.est_admin:
                u.credits += COUT_CREDIT_ETUDE_QUANT
                session.add(MouvementCredit(
                    id=uuid.uuid4().hex[:16], utilisateur_id=u.id, delta=COUT_CREDIT_ETUDE_QUANT,
                    motif="Remboursement — échec de la génération de l'étude quantitative",
                ))
            session.commit()
    finally:
        session.close()


def _etude_quant_vers_dict(e: EtudeQuantitative) -> dict:
    return {
        "id": e.id, "theme": e.theme, "question_recherche": e.question_recherche,
        "langue": e.langue, "statut": e.statut, "etape": e.etape, "contenu": e.contenu, "erreur": e.erreur,
        "modele": e.modele, "cree_le": e.cree_le.isoformat() if e.cree_le else None,
    }


class EtudeQuantIn(BaseModel):
    theme: str
    question_recherche: str = ""
    langue: str = "fr"


@app.get("/api/etudes-quantitatives")
def lister_etudes_quant(user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        etudes = session.query(EtudeQuantitative).filter_by(proprietaire_id=user.id).order_by(EtudeQuantitative.cree_le.desc()).all()
        return [_etude_quant_vers_dict(e) for e in etudes]
    finally:
        session.close()


@app.get("/api/etudes-quantitatives/{etude_id}")
def detail_etude_quant(etude_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        e = session.get(EtudeQuantitative, etude_id)
        if not e or e.proprietaire_id != user.id:
            raise HTTPException(404, "Étude introuvable.")
        return _etude_quant_vers_dict(e)
    finally:
        session.close()


@app.post("/api/etudes-quantitatives")
def creer_etude_quant(payload: EtudeQuantIn, user: Utilisateur = Depends(utilisateur_courant)):
    if not payload.theme.strip():
        raise HTTPException(422, "Le thème de recherche est requis.")
    theme = _nettoyer_texte_utilisateur(payload.theme.strip())
    question_recherche = _nettoyer_texte_utilisateur(payload.question_recherche.strip())
    langue = payload.langue if payload.langue in ("fr", "en") else "fr"
    session = get_session()
    try:
        if not user.email_verifie:
            raise HTTPException(403, "Vérifie ton adresse e-mail avant de générer une étude quantitative.")
        if not user.est_admin and user.credits < COUT_CREDIT_ETUDE_QUANT:
            raise HTTPException(402, "Crédits insuffisants. Consulte les forfaits disponibles dans l'app.")
        if not user.est_admin:
            u = session.get(Utilisateur, user.id)
            u.credits -= COUT_CREDIT_ETUDE_QUANT
            session.add(MouvementCredit(
                id=uuid.uuid4().hex[:16], utilisateur_id=user.id, delta=-COUT_CREDIT_ETUDE_QUANT,
                motif=f"Étude quantitative — {theme[:60]}",
            ))
        etude_id = uuid.uuid4().hex[:16]
        e = EtudeQuantitative(
            id=etude_id, proprietaire_id=user.id, theme=theme,
            question_recherche=question_recherche, langue=langue, statut="en_cours",
        )
        session.add(e)
        session.commit()
    finally:
        session.close()

    threading.Thread(target=_run_etude_quant, args=(etude_id, theme, question_recherche, langue, user.id), daemon=True).start()
    return {"id": etude_id, "statut": "en_cours"}


@app.delete("/api/etudes-quantitatives/{etude_id}")
def supprimer_etude_quant(etude_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        e = session.get(EtudeQuantitative, etude_id)
        if not e or e.proprietaire_id != user.id:
            raise HTTPException(404, "Étude introuvable.")
        session.query(AnalyseQuantitative).filter_by(etude_id=etude_id).delete()
        session.delete(e)
        session.commit()
        return {"ok": True}
    finally:
        session.close()


@app.get("/api/etudes-quantitatives/{etude_id}/export/docx")
def export_docx_etude_quant(etude_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        e = session.get(EtudeQuantitative, etude_id)
        if not e or e.proprietaire_id != user.id:
            raise HTTPException(404, "Étude introuvable.")
        if e.statut != "termine" or not e.contenu:
            raise HTTPException(409, "Cette étude n'est pas encore prête.")
        data = _etude_quant_vers_dict(e)
    finally:
        session.close()
    buf = generer_docx_etude_quant(data)
    nom = _nom_fichier(data["theme"])
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="etude_{nom}.docx"'},
    )


@app.get("/api/etudes-quantitatives/{etude_id}/export/template")
def export_template_etude_quant(etude_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        e = session.get(EtudeQuantitative, etude_id)
        if not e or e.proprietaire_id != user.id:
            raise HTTPException(404, "Étude introuvable.")
        if e.statut != "termine" or not e.contenu:
            raise HTTPException(409, "Cette étude n'est pas encore prête.")
        data = _etude_quant_vers_dict(e)
    finally:
        session.close()
    buf = generer_xlsx_template_questionnaire(data)
    nom = _nom_fichier(data["theme"])
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="gabarit_{nom}.xlsx"'},
    )


def _analyse_quant_vers_dict(a: AnalyseQuantitative) -> dict:
    return {
        "id": a.id, "etude_id": a.etude_id, "nom_fichier": a.nom_fichier, "statut": a.statut,
        "resultats": a.resultats, "erreur": a.erreur, "modele": a.modele,
        "cree_le": a.cree_le.isoformat() if a.cree_le else None,
    }


def _lire_donnees_xlsx(chemin: str, questionnaire: dict) -> list[dict]:
    """Lit le fichier Excel importé par l'utilisateur (rempli à partir du gabarit) et
    le convertit en liste de dicts {code_item: valeur}, une entrée par répondant."""
    from openpyxl import load_workbook
    wb = load_workbook(chemin, data_only=True)
    ws = wb["Réponses"] if "Réponses" in wb.sheetnames else wb.active

    codes_valides = {
        item["code"]
        for section in questionnaire.get("sections", [])
        for item in section.get("items", [])
    }
    entetes = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    lignes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        ligne = {}
        for code, valeur in zip(entetes, row):
            if code in codes_valides and valeur is not None:
                ligne[code] = valeur
        if ligne:
            lignes.append(ligne)
    return lignes


@app.post("/api/etudes-quantitatives/{etude_id}/donnees")
async def importer_donnees_quant(etude_id: str, fichier: UploadFile = File(...), user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        e = session.get(EtudeQuantitative, etude_id)
        if not e or e.proprietaire_id != user.id:
            raise HTTPException(404, "Étude introuvable.")
        if e.statut != "termine" or not e.contenu:
            raise HTTPException(409, "Cette étude n'est pas encore prête.")
        if not user.est_admin and user.credits < COUT_CREDIT_ANALYSE_QUANT:
            raise HTTPException(402, "Crédits insuffisants. Consulte les forfaits disponibles dans l'app.")

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        with tmp as f:
            f.write(await fichier.read())

        try:
            lignes = _lire_donnees_xlsx(tmp.name, e.contenu["questionnaire"])
        except Exception as ex:  # noqa: BLE001
            os.unlink(tmp.name)
            raise HTTPException(422, f"Impossible de lire ce fichier Excel : {ex}")
        os.unlink(tmp.name)

        if len(lignes) < 3:
            raise HTTPException(422, "Au moins 3 répondants sont nécessaires pour une analyse statistique.")

        if not user.est_admin:
            u = session.get(Utilisateur, user.id)
            u.credits -= COUT_CREDIT_ANALYSE_QUANT
            session.add(MouvementCredit(
                id=uuid.uuid4().hex[:16], utilisateur_id=user.id, delta=-COUT_CREDIT_ANALYSE_QUANT,
                motif=f"Analyse quantitative — {e.theme[:60]}",
            ))

        try:
            resultats = analyser_donnees(lignes, e.contenu["questionnaire"])
            statut, erreur = "termine", None
        except Exception as ex:  # noqa: BLE001
            resultats, statut, erreur = None, "erreur", str(ex)
            if not user.est_admin:
                u = session.get(Utilisateur, user.id)
                u.credits += COUT_CREDIT_ANALYSE_QUANT
                session.add(MouvementCredit(
                    id=uuid.uuid4().hex[:16], utilisateur_id=user.id, delta=COUT_CREDIT_ANALYSE_QUANT,
                    motif="Remboursement — échec de l'analyse quantitative",
                ))

        analyse_id = uuid.uuid4().hex[:16]
        a = AnalyseQuantitative(
            id=analyse_id, etude_id=etude_id, proprietaire_id=user.id,
            nom_fichier=fichier.filename or "donnees.xlsx", statut=statut,
            resultats=resultats, erreur=erreur,
        )
        session.add(a)
        session.commit()
        return _analyse_quant_vers_dict(a)
    finally:
        session.close()


@app.get("/api/etudes-quantitatives/{etude_id}/donnees")
def lister_analyses_quant(etude_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        e = session.get(EtudeQuantitative, etude_id)
        if not e or e.proprietaire_id != user.id:
            raise HTTPException(404, "Étude introuvable.")
        analyses = session.query(AnalyseQuantitative).filter_by(etude_id=etude_id).order_by(AnalyseQuantitative.cree_le.desc()).all()
        return [_analyse_quant_vers_dict(a) for a in analyses]
    finally:
        session.close()


@app.get("/api/etudes-quantitatives/{etude_id}/donnees/{analyse_id}/export/docx")
def export_docx_analyse_quant(etude_id: str, analyse_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        e = session.get(EtudeQuantitative, etude_id)
        a = session.get(AnalyseQuantitative, analyse_id)
        if not e or e.proprietaire_id != user.id or not a or a.etude_id != etude_id:
            raise HTTPException(404, "Introuvable.")
        if a.statut != "termine" or not a.resultats:
            raise HTTPException(409, "Cette analyse n'est pas disponible.")
        data_etude, data_analyse = _etude_quant_vers_dict(e), _analyse_quant_vers_dict(a)
    finally:
        session.close()
    buf = generer_docx_analyse_quant(data_etude, data_analyse)
    nom = _nom_fichier(data_etude["theme"])
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="analyse_{nom}.docx"'},
    )


@app.get("/api/etudes-quantitatives/{etude_id}/donnees/{analyse_id}/export/xlsx")
def export_xlsx_analyse_quant(etude_id: str, analyse_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        e = session.get(EtudeQuantitative, etude_id)
        a = session.get(AnalyseQuantitative, analyse_id)
        if not e or e.proprietaire_id != user.id or not a or a.etude_id != etude_id:
            raise HTTPException(404, "Introuvable.")
        if a.statut != "termine" or not a.resultats:
            raise HTTPException(409, "Cette analyse n'est pas disponible.")
        data_etude, data_analyse = _etude_quant_vers_dict(e), _analyse_quant_vers_dict(a)
    finally:
        session.close()
    buf = generer_xlsx_analyse_quant(data_etude, data_analyse)
    nom = _nom_fichier(data_etude["theme"])
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="analyse_{nom}.xlsx"'},
    )


# ----------------------------------------------------------------- authentification
def _generer_code() -> str:
    return f"{secrets.randbelow(1000000):06d}"


def _expire(dt) -> bool:
    """Vrai si la date d'expiration est dépassée. Gère le cas SQLite (dev/test) où les
    dates reviennent sans fuseau horaire, contrairement à PostgreSQL en production."""
    if dt is None:
        return True
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt < datetime.now(timezone.utc)


def _utilisateur_public(u: Utilisateur) -> dict:
    return {
        "id": u.id, "email": u.email, "nom": u.nom,
        "contribution_langues_locales": u.contribution_langues_locales,
        "email_verifie": u.email_verifie,
        "est_admin": u.est_admin,
        "credits": u.credits,
        "forfait_actuel": u.forfait_actuel,
    }


@app.post("/api/auth/inscription")
def inscription(payload: InscriptionIn):
    email = payload.email.strip().lower()
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        raise HTTPException(422, "Adresse e-mail invalide.")
    if len(payload.mot_de_passe) < 8:
        raise HTTPException(422, "Le mot de passe doit contenir au moins 8 caractères.")

    if not payload.accepte_cgu:
        raise HTTPException(422, "Tu dois accepter les conditions d'utilisation pour créer un compte.")

    session = get_session()
    try:
        if session.query(Utilisateur).filter_by(email=email).first():
            raise HTTPException(409, "Un compte existe déjà avec cet e-mail.")

        code = _generer_code()
        user = Utilisateur(
            id=uuid.uuid4().hex[:16], email=email, nom=payload.nom.strip(),
            mot_de_passe_hash=hacher_mot_de_passe(payload.mot_de_passe),
            code_verification=code,
            code_verification_expire=datetime.now(timezone.utc) + timedelta(minutes=30),
            cgu_acceptees_le=datetime.now(timezone.utc), cgu_version=CGU_VERSION,
        )
        if email == ADMIN_EMAIL:
            user.est_admin = True
            user.email_verifie = True
            user.credits = 999999
        elif not EMAIL_CONFIGURE:
            # Pas de serveur SMTP configuré sur Railway : on ne peut pas envoyer de code,
            # donc on ne bloque pas l'inscription — vérification et essai gratuit sont
            # accordés automatiquement. Configure SMTP_HOST/SMTP_USER/SMTP_PASSWORD pour
            # activer la vraie vérification par e-mail.
            user.email_verifie = True
            user.credits = CREDITS_ESSAI_GRATUIT
            user.code_verification = None
            user.code_verification_expire = None
        session.add(user)
        session.flush()
        if user.credits == CREDITS_ESSAI_GRATUIT and email != ADMIN_EMAIL:
            session.add(MouvementCredit(
                id=uuid.uuid4().hex[:16], utilisateur_id=user.id,
                delta=CREDITS_ESSAI_GRATUIT, motif="Essai gratuit à l'inscription",
            ))
        session.commit()

        email_envoye = False
        if EMAIL_CONFIGURE and not user.email_verifie:
            email_envoye = envoyer_code_verification(email, code)
            if not email_envoye:
                # L'envoi a échoué (adresse invalide, quota, etc.) : ne bloque pas non plus,
                # l'utilisateur pourra redemander un code plus tard via /renvoyer-code.
                pass

        return {
            "token": creer_jeton(user.id), "utilisateur": _utilisateur_public(user),
            "email_envoye": email_envoye,
        }
    finally:
        session.close()


@app.post("/api/auth/connexion")
def connexion(payload: ConnexionIn):
    session = get_session()
    try:
        user = session.query(Utilisateur).filter_by(email=payload.email.strip().lower()).first()
        if not user or not verifier_mot_de_passe(payload.mot_de_passe, user.mot_de_passe_hash):
            raise HTTPException(401, "E-mail ou mot de passe incorrect.")
        return {"token": creer_jeton(user.id), "utilisateur": _utilisateur_public(user)}
    finally:
        session.close()


@app.get("/api/auth/moi")
def moi(user: Utilisateur = Depends(utilisateur_courant)):
    return _utilisateur_public(user)


@app.patch("/api/auth/moi")
def modifier_compte(payload: ModifierCompteIn, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        u = session.get(Utilisateur, user.id)
        if payload.nom is not None:
            u.nom = payload.nom.strip()
        if payload.nouveau_mot_de_passe:
            if not payload.mot_de_passe_actuel or not verifier_mot_de_passe(payload.mot_de_passe_actuel, u.mot_de_passe_hash):
                raise HTTPException(401, "Mot de passe actuel incorrect.")
            if len(payload.nouveau_mot_de_passe) < 8:
                raise HTTPException(422, "Le nouveau mot de passe doit contenir au moins 8 caractères.")
            u.mot_de_passe_hash = hacher_mot_de_passe(payload.nouveau_mot_de_passe)
        session.commit()
        return _utilisateur_public(u)
    finally:
        session.close()


def _supprimer_utilisateur_cascade(session, user_id: str):
    """Supprime un compte et ses données associées. Les entretiens des autres membres
    d'un corpus que ce compte possédait sont conservés, mais détachés du corpus."""
    corpus_possedes = session.query(Corpus).filter_by(proprietaire_id=user_id).all()
    for c in corpus_possedes:
        session.query(Entretien).filter_by(corpus_id=c.id).update({"corpus_id": None})
        session.query(MembreCorpus).filter_by(corpus_id=c.id).delete()
        session.delete(c)

    entretiens_possedes = session.query(Entretien).filter_by(proprietaire_id=user_id).all()
    for e in entretiens_possedes:
        session.query(Codage).filter_by(entretien_id=e.id).delete()
        session.delete(e)

    session.query(MembreCorpus).filter_by(utilisateur_id=user_id).delete()
    session.query(Codage).filter_by(utilisateur_id=user_id).delete()
    session.query(ContributionLangue).filter_by(utilisateur_id=user_id).delete()
    session.query(MouvementCredit).filter_by(utilisateur_id=user_id).delete()
    session.delete(session.get(Utilisateur, user_id))


@app.delete("/api/auth/moi")
def supprimer_compte(payload: SupprimerCompteIn, user: Utilisateur = Depends(utilisateur_courant)):
    if user.email == ADMIN_EMAIL:
        raise HTTPException(403, "Le compte administrateur ne peut pas être auto-supprimé.")
    session = get_session()
    try:
        u = session.get(Utilisateur, user.id)
        if not verifier_mot_de_passe(payload.mot_de_passe, u.mot_de_passe_hash):
            raise HTTPException(401, "Mot de passe incorrect.")
        _supprimer_utilisateur_cascade(session, user.id)
        session.commit()
        return {"ok": True}
    finally:
        session.close()


@app.post("/api/auth/verifier-email")
def verifier_email(payload: VerifierEmailIn):
    session = get_session()
    try:
        user = session.query(Utilisateur).filter_by(email=payload.email.strip().lower()).first()
        if not user:
            raise HTTPException(404, "Compte introuvable.")
        if user.email_verifie:
            return {"email_verifie": True, "credits": user.credits}
        if (
            not user.code_verification
            or user.code_verification != payload.code.strip()
            or _expire(user.code_verification_expire)
        ):
            raise HTTPException(400, "Code invalide ou expiré.")

        user.email_verifie = True
        user.code_verification = None
        user.code_verification_expire = None
        user.credits += CREDITS_ESSAI_GRATUIT
        session.add(MouvementCredit(
            id=uuid.uuid4().hex[:16], utilisateur_id=user.id,
            delta=CREDITS_ESSAI_GRATUIT, motif="Essai gratuit à l'inscription",
        ))
        session.commit()
        return {"email_verifie": True, "credits": user.credits}
    finally:
        session.close()


@app.post("/api/auth/renvoyer-code")
def renvoyer_code(payload: RenvoyerCodeIn):
    session = get_session()
    try:
        user = session.query(Utilisateur).filter_by(email=payload.email.strip().lower()).first()
        if not user or user.email_verifie:
            return {"ok": True}  # réponse neutre : n'indique jamais si le compte existe
        code = _generer_code()
        user.code_verification = code
        user.code_verification_expire = datetime.now(timezone.utc) + timedelta(minutes=30)
        session.commit()
        envoye = envoyer_code_verification(user.email, code) if EMAIL_CONFIGURE else False
        return {"ok": True, "email_envoye": envoye}
    finally:
        session.close()


@app.post("/api/auth/mot-de-passe-oublie")
def mot_de_passe_oublie(payload: MotDePasseOublieIn):
    session = get_session()
    try:
        user = session.query(Utilisateur).filter_by(email=payload.email.strip().lower()).first()
        if user:
            code = _generer_code()
            user.code_reinitialisation = code
            user.code_reinitialisation_expire = datetime.now(timezone.utc) + timedelta(minutes=30)
            session.commit()
            if EMAIL_CONFIGURE:
                envoyer_code_reinitialisation(user.email, code)
        # Toujours la même réponse, que le compte existe ou non — évite de révéler
        # quelles adresses sont enregistrées.
        return {"ok": True}
    finally:
        session.close()


@app.post("/api/auth/reinitialiser-mot-de-passe")
def reinitialiser_mot_de_passe(payload: ReinitialiserMdpIn):
    if len(payload.nouveau_mot_de_passe) < 8:
        raise HTTPException(422, "Le mot de passe doit contenir au moins 8 caractères.")
    session = get_session()
    try:
        user = session.query(Utilisateur).filter_by(email=payload.email.strip().lower()).first()
        if (
            not user or not user.code_reinitialisation
            or user.code_reinitialisation != payload.code.strip()
            or _expire(user.code_reinitialisation_expire)
        ):
            raise HTTPException(400, "Code invalide ou expiré.")
        user.mot_de_passe_hash = hacher_mot_de_passe(payload.nouveau_mot_de_passe)
        user.code_reinitialisation = None
        user.code_reinitialisation_expire = None
        session.commit()
        return {"token": creer_jeton(user.id), "utilisateur": _utilisateur_public(user)}
    finally:
        session.close()


@app.post("/api/auth/preferences")
def maj_preferences(payload: PreferencesIn, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        u = session.get(Utilisateur, user.id)
        u.contribution_langues_locales = payload.contribution_langues_locales
        session.commit()
        return {"contribution_langues_locales": u.contribution_langues_locales}
    finally:
        session.close()


@app.post("/api/contributions")
def enregistrer_contribution(payload: ContributionIn, user: Utilisateur = Depends(utilisateur_courant)):
    if payload.langue not in ("dyu", "bci"):
        raise HTTPException(422, "Seules les langues locales expérimentales acceptent des contributions.")
    session = get_session()
    try:
        u = session.get(Utilisateur, user.id)
        if not u.contribution_langues_locales:
            raise HTTPException(403, "Active d'abord la contribution dans tes réglages.")
        session.add(ContributionLangue(
            id=uuid.uuid4().hex[:16], utilisateur_id=user.id, langue=payload.langue,
            texte_original=payload.texte_original, texte_corrige=payload.texte_corrige,
        ))
        session.commit()
        return {"ok": True}
    finally:
        session.close()


@app.get("/api/contributions/nombre")
def nombre_contributions(user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        n = session.query(ContributionLangue).filter_by(utilisateur_id=user.id).count()
        return {"nombre": n}
    finally:
        session.close()


# ----------------------------------------------------------------- administration
def admin_requis(user: Utilisateur = Depends(utilisateur_courant)) -> Utilisateur:
    if not user.est_admin:
        raise HTTPException(403, "Accès réservé à l'administrateur.")
    return user


@app.get("/api/admin/utilisateurs")
def admin_lister_utilisateurs(_: Utilisateur = Depends(admin_requis)):
    session = get_session()
    try:
        users = session.query(Utilisateur).order_by(Utilisateur.cree_le.desc()).all()
        return [
            {
                "id": u.id, "email": u.email, "nom": u.nom, "email_verifie": u.email_verifie,
                "est_admin": u.est_admin, "credits": u.credits, "forfait_actuel": u.forfait_actuel,
                "cree_le": u.cree_le.isoformat() if u.cree_le else None,
            }
            for u in users
        ]
    finally:
        session.close()


@app.post("/api/admin/utilisateurs/{utilisateur_id}/credits")
def admin_ajuster_credits(utilisateur_id: str, payload: CreditsIn, admin: Utilisateur = Depends(admin_requis)):
    session = get_session()
    try:
        u = session.get(Utilisateur, utilisateur_id)
        if not u:
            raise HTTPException(404, "Utilisateur introuvable.")
        u.credits += payload.delta
        session.add(MouvementCredit(
            id=uuid.uuid4().hex[:16], utilisateur_id=u.id, delta=payload.delta,
            motif=payload.motif or f"Ajustement manuel par {admin.email}",
        ))
        session.commit()
        return {"credits": u.credits}
    finally:
        session.close()


@app.post("/api/admin/utilisateurs/{utilisateur_id}/credits/definir")
def admin_definir_credits(utilisateur_id: str, payload: DefinirCreditsIn, admin: Utilisateur = Depends(admin_requis)):
    session = get_session()
    try:
        u = session.get(Utilisateur, utilisateur_id)
        if not u:
            raise HTTPException(404, "Utilisateur introuvable.")
        delta = payload.valeur - u.credits
        u.credits = payload.valeur
        session.add(MouvementCredit(
            id=uuid.uuid4().hex[:16], utilisateur_id=u.id, delta=delta,
            motif=payload.motif or f"Solde redéfini par {admin.email}",
        ))
        session.commit()
        return {"credits": u.credits}
    finally:
        session.close()


@app.delete("/api/admin/utilisateurs/{utilisateur_id}")
def admin_supprimer_utilisateur(utilisateur_id: str, admin: Utilisateur = Depends(admin_requis)):
    session = get_session()
    try:
        u = session.get(Utilisateur, utilisateur_id)
        if not u:
            raise HTTPException(404, "Utilisateur introuvable.")
        if u.email == ADMIN_EMAIL:
            raise HTTPException(403, "Le compte administrateur ne peut pas être supprimé.")
        _supprimer_utilisateur_cascade(session, utilisateur_id)
        session.commit()
        return {"ok": True}
    finally:
        session.close()


@app.get("/api/admin/utilisateurs/{utilisateur_id}/mouvements")
def admin_mouvements_credit(utilisateur_id: str, _: Utilisateur = Depends(admin_requis)):
    session = get_session()
    try:
        mvts = session.query(MouvementCredit).filter_by(utilisateur_id=utilisateur_id) \
            .order_by(MouvementCredit.cree_le.desc()).all()
        return [{"delta": m.delta, "motif": m.motif, "cree_le": m.cree_le.isoformat()} for m in mvts]
    finally:
        session.close()


@app.post("/api/admin/utilisateurs/{utilisateur_id}/attribuer-forfait")
def admin_attribuer_forfait(utilisateur_id: str, payload: AttribuerForfaitIn, admin: Utilisateur = Depends(admin_requis)):
    session = get_session()
    try:
        u = session.get(Utilisateur, utilisateur_id)
        forfait = session.get(Forfait, payload.forfait_id)
        if not u or not forfait:
            raise HTTPException(404, "Utilisateur ou forfait introuvable.")
        u.credits += forfait.credits_inclus
        u.forfait_actuel = forfait.nom
        session.add(MouvementCredit(
            id=uuid.uuid4().hex[:16], utilisateur_id=u.id, delta=forfait.credits_inclus,
            motif=f"Forfait « {forfait.nom} » attribué par {admin.email}",
        ))
        session.commit()
        return {"credits": u.credits, "forfait_actuel": u.forfait_actuel}
    finally:
        session.close()


@app.get("/api/forfaits")
def lister_forfaits():
    """Catalogue public des forfaits actifs, affiché dans l'app."""
    session = get_session()
    try:
        forfaits = session.query(Forfait).filter_by(actif=True).order_by(Forfait.prix_fcfa).all()
        return [
            {"id": f.id, "nom": f.nom, "prix_fcfa": f.prix_fcfa, "credits_inclus": f.credits_inclus, "description": f.description}
            for f in forfaits
        ]
    finally:
        session.close()


# ----------------------------------------------------------------- recharge de crédits à la carte
@app.get("/api/recharges/tarif")
def tarif_recharge():
    """Grille tarifaire publique de la recharge à la carte, affichée dans l'app avant paiement."""
    disponible = _paydunya_pret() if PAYMENT_PROVIDER == "paydunya" else bool(PAYMENT_PROVIDER)
    return {
        "prix_credit_fcfa": PRIX_CREDIT_FCFA, "credits_min": CREDITS_MIN_RECHARGE,
        "paiement_disponible": disponible,
    }


def _commande_vers_dict(c: Commande) -> dict:
    return {
        "id": c.id, "credits": c.credits, "montant_fcfa": c.montant_fcfa, "statut": c.statut,
        "lien_paiement": c.lien_paiement, "cree_le": c.cree_le.isoformat() if c.cree_le else None,
    }


def _initier_paiement_fournisseur(commande: Commande, utilisateur: Utilisateur) -> str:
    """Point d'intégration avec l'agrégateur de paiement. PayDunya est implémenté ;
    d'autres fournisseurs pourront être ajoutés selon le même principe."""
    if PAYMENT_PROVIDER == "paydunya":
        if not (PAYDUNYA_MASTER_KEY and PAYDUNYA_PRIVATE_KEY and PAYDUNYA_TOKEN):
            raise NotImplementedError(
                "PayDunya n'est pas complètement configuré : PAYDUNYA_PRIVATE_KEY et/ou "
                "PAYDUNYA_TOKEN manquent dans les variables Railway (la clé maîtresse seule "
                "ne suffit jamais à authentifier une requête PayDunya)."
            )
        if not BACKEND_PUBLIC_URL:
            raise NotImplementedError(
                "BACKEND_PUBLIC_URL doit être configurée (adresse publique du serveur Railway, "
                "utilisée pour recevoir la confirmation de paiement)."
            )
        payload = {
            "invoice": {
                "total_amount": commande.montant_fcfa,
                "description": f"Djeliya — {commande.credits:g} crédits",
            },
            "store": {"name": "Djeliya"},
            "actions": {
                "callback_url": f"{BACKEND_PUBLIC_URL}/api/recharges/webhook/paydunya",
                "return_url": BACKEND_PUBLIC_URL,
            },
            "custom_data": {"commande_id": commande.id},
        }
        try:
            resp = requests.post(
                f"{_paydunya_base_url()}/checkout-invoice/create",
                json=payload, headers=_paydunya_headers(), timeout=15,
            )
            data = resp.json()
        except Exception as ex:  # noqa: BLE001
            raise NotImplementedError(f"Erreur de connexion à PayDunya : {ex}")

        if data.get("response_code") != "00":
            raise NotImplementedError(
                f"PayDunya a refusé la création de la facture : {data.get('response_text') or data}"
            )
        commande.reference_fournisseur = data["token"]
        return data["response_text"]  # URL de paiement à ouvrir dans l'app

    raise HTTPException(
        501,
        "Le paiement en ligne n'est pas encore configuré sur ce serveur "
        "(variable PAYMENT_PROVIDER absente). Contacte l'administrateur.",
    )


@app.post("/api/recharges")
def creer_recharge(payload: RechargeIn, user: Utilisateur = Depends(utilisateur_courant)):
    if payload.credits < CREDITS_MIN_RECHARGE:
        raise HTTPException(422, f"Le minimum est de {CREDITS_MIN_RECHARGE:g} crédits par recharge.")
    if not user.email_verifie:
        raise HTTPException(403, "Vérifie ton adresse e-mail avant d'effectuer un paiement.")

    montant = round(payload.credits * PRIX_CREDIT_FCFA)
    session = get_session()
    try:
        commande = Commande(
            id=uuid.uuid4().hex[:16], utilisateur_id=user.id, credits=payload.credits,
            montant_fcfa=montant, statut="en_attente", fournisseur=PAYMENT_PROVIDER,
        )
        session.add(commande)
        session.commit()

        try:
            lien = _initier_paiement_fournisseur(commande, user)
            commande.lien_paiement = lien
            session.commit()
        except HTTPException:
            session.delete(commande)
            session.commit()
            raise
        except NotImplementedError as ex:
            session.delete(commande)
            session.commit()
            raise HTTPException(501, str(ex))

        return _commande_vers_dict(commande)
    finally:
        session.close()


@app.get("/api/recharges")
def lister_recharges(user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        commandes = session.query(Commande).filter_by(utilisateur_id=user.id).order_by(Commande.cree_le.desc()).all()
        return [_commande_vers_dict(c) for c in commandes]
    finally:
        session.close()


@app.get("/api/recharges/{commande_id}")
def detail_recharge(commande_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        c = session.get(Commande, commande_id)
        if not c or c.utilisateur_id != user.id:
            raise HTTPException(404, "Commande introuvable.")
        return _commande_vers_dict(c)
    finally:
        session.close()


def _verifier_statut_paydunya(token: str) -> dict:
    """Interroge directement l'API PayDunya pour connaître le VRAI statut d'une
    facture, indépendamment de tout webhook — status : completed | pending |
    canceled | failed, selon la documentation officielle."""
    resp = requests.get(
        f"{_paydunya_base_url()}/checkout-invoice/confirm/{token}",
        headers=_paydunya_headers(), timeout=15,
    )
    return resp.json()


def _confirmer_et_crediter(session, commande_id: str) -> Commande:
    """Marque une commande payée et crédite le compte, de façon idempotente —
    utilisée à la fois par le webhook et par la vérification manuelle, pour ne
    jamais avoir deux logiques de crédit divergentes."""
    commande = session.get(Commande, commande_id)
    if not commande or commande.statut == "payee":
        return commande
    commande.statut = "payee"
    commande.payee_le = datetime.now(timezone.utc)
    u = session.get(Utilisateur, commande.utilisateur_id)
    u.credits += commande.credits
    session.add(MouvementCredit(
        id=uuid.uuid4().hex[:16], utilisateur_id=u.id, delta=commande.credits,
        motif=f"Recharge payée — {commande.montant_fcfa} FCFA via PayDunya",
    ))
    session.commit()
    return commande


@app.post("/api/recharges/{commande_id}/verifier")
def verifier_recharge_manuellement(commande_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    """Permet au chercheur (ou à l'app, automatiquement) de forcer une nouvelle
    vérification du statut réel auprès de PayDunya — utile si le webhook n'est
    jamais arrivé (pare-feu, latence réseau...) alors que le paiement a bien
    été effectué côté PayDunya."""
    session = get_session()
    try:
        commande = session.get(Commande, commande_id)
        if not commande or commande.utilisateur_id != user.id:
            raise HTTPException(404, "Commande introuvable.")
        if commande.statut == "payee":
            return _commande_vers_dict(commande)
        if not commande.reference_fournisseur:
            raise HTTPException(409, "Cette commande n'a pas de référence de paiement à vérifier.")

        try:
            verif = _verifier_statut_paydunya(commande.reference_fournisseur)
        except Exception:  # noqa: BLE001
            raise HTTPException(502, "Impossible de joindre PayDunya pour vérifier ce paiement.")

        statut_paydunya = verif.get("status")
        if statut_paydunya == "completed":
            commande = _confirmer_et_crediter(session, commande_id)
        elif statut_paydunya in ("canceled", "fail"):
            commande.statut = "echouee"
            session.commit()
        return {**_commande_vers_dict(commande), "statut_paydunya_brut": statut_paydunya}
    finally:
        session.close()


@app.post("/api/recharges/webhook/{fournisseur}")
async def webhook_paiement(fournisseur: str, request: Request):
    """Point de confirmation de paiement appelé par le fournisseur (serveur à serveur).

    Sécurité à deux niveaux, jamais l'un sans l'autre :
    1. Vérification de la signature du webhook (hash SHA-512 de la clé maîtresse,
       tel que documenté par PayDunya) — rejette toute requête forgée.
    2. Re-vérification indépendante du statut du paiement directement auprès de
       l'API PayDunya (jamais confiance aveugle au contenu du webhook, même signé).
    """
    if fournisseur != "paydunya":
        raise HTTPException(501, f"Webhook « {fournisseur} » non configuré sur ce serveur.")
    if not PAYDUNYA_MASTER_KEY:
        raise HTTPException(501, "PayDunya n'est pas configuré sur ce serveur.")

    contenu_type = request.headers.get("content-type", "")
    try:
        if "application/json" in contenu_type:
            charge = await request.json()
        else:
            form = await request.form()
            charge = json.loads(form.get("data", "{}"))
    except Exception:  # noqa: BLE001
        raise HTTPException(400, "Corps de requête invalide.")

    hash_attendu = hashlib.sha512(PAYDUNYA_MASTER_KEY.encode()).hexdigest()
    if charge.get("hash") != hash_attendu:
        raise HTTPException(403, "Signature invalide — requête rejetée.")

    commande_id = (charge.get("custom_data") or {}).get("commande_id")
    token = (charge.get("invoice") or {}).get("token") or charge.get("token")
    if not commande_id or not token:
        raise HTTPException(400, "Données de webhook incomplètes.")

    try:
        verif = _verifier_statut_paydunya(token)
    except Exception:  # noqa: BLE001
        raise HTTPException(502, "Impossible de vérifier le paiement auprès de PayDunya.")

    if verif.get("status") != "completed":
        return {"ok": True, "traite": False}

    session = get_session()
    try:
        _confirmer_et_crediter(session, commande_id)
        return {"ok": True, "traite": True}
    finally:
        session.close()


@app.post("/api/admin/forfaits")
def admin_creer_forfait(payload: ForfaitIn, _: Utilisateur = Depends(admin_requis)):
    session = get_session()
    try:
        f = Forfait(
            id=uuid.uuid4().hex[:16], nom=payload.nom, prix_fcfa=payload.prix_fcfa,
            credits_inclus=payload.credits_inclus, description=payload.description,
        )
        session.add(f)
        session.commit()
        return {"id": f.id}
    finally:
        session.close()


@app.post("/api/admin/forfaits/{forfait_id}/desactiver")
def admin_desactiver_forfait(forfait_id: str, _: Utilisateur = Depends(admin_requis)):
    session = get_session()
    try:
        f = session.get(Forfait, forfait_id)
        if not f:
            raise HTTPException(404, "Forfait introuvable.")
        f.actif = False
        session.commit()
        return {"ok": True}
    finally:
        session.close()
# ----------------------------------------------------------------- corpus / équipes
@app.post("/api/corpus")
def creer_corpus(payload: CorpusIn, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        corpus = Corpus(
            id=uuid.uuid4().hex[:16], nom=payload.nom.strip() or "Corpus sans nom",
            proprietaire_id=user.id, code_invitation=_generer_code_invitation(),
        )
        session.add(corpus)
        session.add(MembreCorpus(
            id=uuid.uuid4().hex[:16], corpus_id=corpus.id,
            utilisateur_id=user.id, role="proprietaire",
        ))
        session.commit()
        return {"id": corpus.id, "nom": corpus.nom, "code_invitation": corpus.code_invitation}
    finally:
        session.close()


@app.get("/api/corpus")
def lister_corpus(user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        memberships = session.query(MembreCorpus).filter_by(utilisateur_id=user.id).all()
        out = []
        for m in memberships:
            corpus = session.get(Corpus, m.corpus_id)
            if not corpus:
                continue
            nb_membres = session.query(MembreCorpus).filter_by(corpus_id=corpus.id).count()
            out.append({
                "id": corpus.id, "nom": corpus.nom, "role": m.role,
                "code_invitation": corpus.code_invitation if m.role == "proprietaire" else None,
                "nb_membres": nb_membres,
            })
        return out
    finally:
        session.close()


@app.post("/api/corpus/rejoindre")
def rejoindre_corpus(payload: RejoindreIn, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        corpus = session.query(Corpus).filter_by(code_invitation=payload.code.strip().upper()).first()
        if not corpus:
            raise HTTPException(404, "Code d'invitation invalide.")
        existant = session.query(MembreCorpus).filter_by(corpus_id=corpus.id, utilisateur_id=user.id).first()
        if not existant:
            session.add(MembreCorpus(
                id=uuid.uuid4().hex[:16], corpus_id=corpus.id,
                utilisateur_id=user.id, role="codeur",
            ))
            session.commit()
        return {"id": corpus.id, "nom": corpus.nom}
    finally:
        session.close()


@app.get("/api/corpus/{corpus_id}/membres")
def membres_corpus(corpus_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        if not session.query(MembreCorpus).filter_by(corpus_id=corpus_id, utilisateur_id=user.id).first():
            raise HTTPException(403, "Tu n'as pas accès à ce corpus.")
        membres = session.query(MembreCorpus).filter_by(corpus_id=corpus_id).all()
        out = []
        for m in membres:
            u = session.get(Utilisateur, m.utilisateur_id)
            out.append({"nom": u.nom or u.email, "role": m.role})
        return out
    finally:
        session.close()


def _corpus_vers_dict(c: Corpus) -> dict:
    return {
        "id": c.id, "nom": c.nom, "code_invitation": c.code_invitation,
        "analyse_statut": c.analyse_statut, "analyse": c.analyse, "analyse_erreur": c.analyse_erreur,
        "analyse_contexte": c.analyse_contexte, "analyse_methode": c.analyse_methode,
        "analyse_modele": c.analyse_modele, "analyse_nb_entretiens": c.analyse_nb_entretiens,
        "analyse_langue": c.analyse_langue,
    }


@app.get("/api/corpus/{corpus_id}")
def detail_corpus(corpus_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        if not session.query(MembreCorpus).filter_by(corpus_id=corpus_id, utilisateur_id=user.id).first():
            raise HTTPException(403, "Tu n'as pas accès à ce corpus.")
        c = session.get(Corpus, corpus_id)
        if not c:
            raise HTTPException(404, "Corpus introuvable.")
        return _corpus_vers_dict(c)
    finally:
        session.close()


@app.post("/api/corpus/{corpus_id}/analyser")
def lancer_analyse_corpus(corpus_id: str, payload: AnalyseCorpusIn, user: Utilisateur = Depends(utilisateur_courant)):
    if payload.methode not in METHODES_ANALYSE:
        raise HTTPException(422, f"Méthode inconnue : {payload.methode}")
    langue = payload.langue if payload.langue in ("fr", "en") else "fr"
    session = get_session()
    try:
        if not session.query(MembreCorpus).filter_by(corpus_id=corpus_id, utilisateur_id=user.id).first():
            raise HTTPException(403, "Tu n'as pas accès à ce corpus.")
        c = session.get(Corpus, corpus_id)
        if not c:
            raise HTTPException(404, "Corpus introuvable.")
        nb_termines = session.query(Entretien).filter_by(corpus_id=corpus_id, statut="termine").count()
        if nb_termines < 2:
            raise HTTPException(422, "Il faut au moins 2 entretiens terminés dans ce corpus pour une analyse transversale.")
        cout = COUT_CREDIT_ANALYSE * nb_termines
        if not user.est_admin and user.credits < cout:
            raise HTTPException(402, f"Crédits insuffisants ({cout} nécessaires pour {nb_termines} entretiens).")
        if not user.est_admin:
            u = session.get(Utilisateur, user.id)
            u.credits -= cout
            session.add(MouvementCredit(
                id=uuid.uuid4().hex[:16], utilisateur_id=user.id, delta=-cout,
                motif=f"Analyse transversale ({payload.methode}) — {c.nom}",
            ))
        c.analyse_statut = "en_cours"
        c.analyse_erreur = None
        session.commit()
    finally:
        session.close()

    threading.Thread(target=_run_analyse_corpus, args=(corpus_id, payload.contexte, payload.methode, user.id, cout, langue), daemon=True).start()
    return {"analyse_statut": "en_cours"}


# ----------------------------------------------------------------- transcriptions
def _finaliser_transcription(user: Utilisateur, tmp_path: str, langue: str, vocabulaire: str, corpus_id: str, titre: str, nom_fichier: str) -> dict:
    """Valide, débite les crédits et lance la transcription — logique partagée entre
    l'envoi direct (petits fichiers) et l'envoi fractionné (fichiers longs, > 5 min
    d'envoi sur mobile, pour respecter la limite de requête HTTP de Railway)."""
    entretien_id = uuid.uuid4().hex[:12]
    session = get_session()
    try:
        if not user.email_verifie:
            os.unlink(tmp_path)
            raise HTTPException(403, "Vérifie ton adresse e-mail avant de lancer une transcription.")
        if not user.est_admin and user.credits < COUT_CREDIT_TRANSCRIPTION:
            os.unlink(tmp_path)
            raise HTTPException(402, "Crédits insuffisants. Consulte les forfaits disponibles dans l'app.")
        if corpus_id and not session.query(MembreCorpus).filter_by(corpus_id=corpus_id, utilisateur_id=user.id).first():
            os.unlink(tmp_path)
            raise HTTPException(403, "Tu n'as pas accès à ce corpus.")

        if not user.est_admin:
            u = session.get(Utilisateur, user.id)
            u.credits -= COUT_CREDIT_TRANSCRIPTION
            session.add(MouvementCredit(
                id=uuid.uuid4().hex[:16], utilisateur_id=user.id, delta=-COUT_CREDIT_TRANSCRIPTION,
                motif=f"Transcription — {titre or nom_fichier}",
            ))

        e = Entretien(
            id=entretien_id, proprietaire_id=user.id, corpus_id=corpus_id or None,
            titre=titre or nom_fichier or "Entretien sans titre", langue=langue, statut="en_attente",
        )
        session.add(e)
        session.commit()
    finally:
        session.close()

    threading.Thread(target=_run_job, args=(entretien_id, tmp_path, langue, vocabulaire), daemon=True).start()
    return {"id": entretien_id, "statut": "en_attente"}


@app.post("/api/transcriptions")
async def creer_transcription(
    audio: UploadFile = File(...),
    langue: str = Form("auto"),
    vocabulaire: str = Form(""),
    corpus_id: str = Form(""),
    titre: str = Form(""),
    user: Utilisateur = Depends(utilisateur_courant),
):
    if langue not in LANGUES_SUPPORTEES:
        raise HTTPException(422, f"Langue inconnue : {langue}")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(audio.filename or "a.wav")[1])
    size = 0
    with tmp as f:
        while chunk := await audio.read(1024 * 1024):
            size += len(chunk)
            if size > MAX_UPLOAD_MB * 1024 * 1024:
                os.unlink(tmp.name)
                raise HTTPException(413, f"Fichier trop volumineux (max {MAX_UPLOAD_MB} Mo)")
            f.write(chunk)

    return _finaliser_transcription(user, tmp.name, langue, vocabulaire, corpus_id, titre, audio.filename)


# ------------------------------------------------------------- envoi fractionné (fichiers longs)
# Railway impose une limite plate-forme de 5 minutes par requête HTTP publique, non
# modifiable. Un enregistrement de 1h+ envoyé en une seule requête sur une connexion
# mobile instable peut dépasser ce délai et être coupé en plein transfert (l'app
# affiche alors "Failed to fetch" — la coupure vient du réseau, pas du serveur).
# On découpe donc l'envoi en petits morceaux (quelques Mo chacun, bien sous la limite),
# assemblés côté serveur une fois tous reçus.
SESSIONS_ENVOI: dict[str, dict] = {}


class InitEnvoiIn(BaseModel):
    nom_fichier: str = ""
    langue: str = "auto"
    vocabulaire: str = ""
    corpus_id: str = ""
    titre: str = ""


@app.post("/api/transcriptions/envoi/init")
def init_envoi_fractionne(payload: InitEnvoiIn, user: Utilisateur = Depends(utilisateur_courant)):
    if payload.langue not in LANGUES_SUPPORTEES:
        raise HTTPException(422, f"Langue inconnue : {payload.langue}")
    session_id = uuid.uuid4().hex[:20]
    dossier = os.path.join(tempfile.gettempdir(), f"envoi_{session_id}")
    os.makedirs(dossier, exist_ok=True)
    SESSIONS_ENVOI[session_id] = {
        "utilisateur_id": user.id, "dossier": dossier,
        "nom_fichier": payload.nom_fichier, "langue": payload.langue,
        "vocabulaire": payload.vocabulaire, "corpus_id": payload.corpus_id, "titre": payload.titre,
        "taille": 0,
    }
    return {"session_id": session_id}


@app.post("/api/transcriptions/envoi/{session_id}/morceau")
async def envoyer_morceau(session_id: str, index: int = Form(...), morceau: UploadFile = File(...), user: Utilisateur = Depends(utilisateur_courant)):
    s = SESSIONS_ENVOI.get(session_id)
    if not s or s["utilisateur_id"] != user.id:
        raise HTTPException(404, "Session d'envoi introuvable ou expirée — recommence l'envoi.")
    chemin_morceau = os.path.join(s["dossier"], f"{index:06d}.part")
    taille = 0
    with open(chemin_morceau, "wb") as f:
        while chunk := await morceau.read(1024 * 1024):
            taille += len(chunk)
            if s["taille"] + taille > MAX_UPLOAD_MB * 1024 * 1024:
                raise HTTPException(413, f"Fichier trop volumineux (max {MAX_UPLOAD_MB} Mo)")
            f.write(chunk)
    s["taille"] += taille
    return {"ok": True, "index": index}


@app.post("/api/transcriptions/envoi/{session_id}/terminer")
def terminer_envoi_fractionne(session_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    s = SESSIONS_ENVOI.pop(session_id, None)
    if not s or s["utilisateur_id"] != user.id:
        raise HTTPException(404, "Session d'envoi introuvable ou expirée — recommence l'envoi.")

    morceaux = sorted(os.listdir(s["dossier"]))
    if not morceaux:
        raise HTTPException(422, "Aucun morceau reçu pour cet envoi.")

    suffixe = os.path.splitext(s["nom_fichier"] or "a.wav")[1]
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffixe)
    with tmp as f:
        for nom in morceaux:
            with open(os.path.join(s["dossier"], nom), "rb") as morceau:
                f.write(morceau.read())
    for nom in morceaux:
        os.unlink(os.path.join(s["dossier"], nom))
    os.rmdir(s["dossier"])

    return _finaliser_transcription(
        user, tmp.name, s["langue"], s["vocabulaire"], s["corpus_id"], s["titre"], s["nom_fichier"],
    )


@app.get("/api/transcriptions")
def lister_transcriptions(user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        corpus_ids = [m.corpus_id for m in session.query(MembreCorpus).filter_by(utilisateur_id=user.id).all()]
        q = session.query(Entretien).filter(
            (Entretien.proprietaire_id == user.id) | (Entretien.corpus_id.in_(corpus_ids or ["_aucun_"]))
        )
        return [
            {"id": e.id, "titre": e.titre, "statut": e.statut, "corpus_id": e.corpus_id,
             "cree_le": e.cree_le.isoformat() if e.cree_le else None}
            for e in q.all()
        ]
    finally:
        session.close()


@app.get("/api/transcriptions/{entretien_id}")
def obtenir_transcription(entretien_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        e = session.get(Entretien, entretien_id)
        if not e:
            raise HTTPException(404, "Entretien introuvable.")
        _verifier_acces(session, e, user)
        return _entretien_vers_dict(e)
    finally:
        session.close()


@app.post("/api/transcriptions/{entretien_id}/segments/{index}")
def corriger_segment(entretien_id: str, index: int, payload: dict, user: Utilisateur = Depends(utilisateur_courant)):
    """Corrige UN SEUL segment de façon atomique (jamais tout le tableau d'un coup),
    pour que deux corrections rapprochées ne puissent jamais s'écraser l'une l'autre."""
    session = get_session()
    try:
        e = session.get(Entretien, entretien_id)
        if not e:
            raise HTTPException(404, "Entretien introuvable.")
        _verifier_acces(session, e, user)
        segments = list(e.segments or [])
        if index < 0 or index >= len(segments):
            raise HTTPException(422, "Segment introuvable.")
        segments[index] = payload.get("segment", segments[index])
        e.segments = segments
        session.commit()
        return {"ok": True}
    finally:
        session.close()


@app.get("/api/transcriptions/{entretien_id}/export/docx")
def export_docx_entretien(entretien_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        e = session.get(Entretien, entretien_id)
        if not e:
            raise HTTPException(404, "Entretien introuvable.")
        _verifier_acces(session, e, user)
        data = _entretien_vers_dict(e)
    finally:
        session.close()
    buf = generer_docx_entretien(data)
    nom = _nom_fichier(data["titre"])
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{nom}.docx"'},
    )


@app.get("/api/transcriptions/{entretien_id}/export/xlsx")
def export_xlsx_entretien(entretien_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        e = session.get(Entretien, entretien_id)
        if not e:
            raise HTTPException(404, "Entretien introuvable.")
        _verifier_acces(session, e, user)
        data = _entretien_vers_dict(e)
    finally:
        session.close()
    buf = generer_xlsx_entretien(data)
    nom = _nom_fichier(data["titre"])
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nom}.xlsx"'},
    )


@app.get("/api/corpus/{corpus_id}/export/docx")
def export_docx_corpus(corpus_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        if not session.query(MembreCorpus).filter_by(corpus_id=corpus_id, utilisateur_id=user.id).first():
            raise HTTPException(403, "Tu n'as pas accès à ce corpus.")
        c = session.get(Corpus, corpus_id)
        if not c or not c.analyse:
            raise HTTPException(404, "Aucune analyse de corpus disponible à exporter.")
        entretiens = [_entretien_vers_dict(e) for e in session.query(Entretien).filter_by(corpus_id=corpus_id, statut="termine").order_by(Entretien.cree_le).all()]
        fiabilite = _fiabilite_corpus(session, corpus_id)
        buf = generer_docx_etude(_corpus_vers_dict(c), entretiens, fiabilite)
        nom = _nom_fichier(c.nom)
    finally:
        session.close()
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="rapport_{nom}.docx"'},
    )


@app.get("/api/corpus/{corpus_id}/export/xlsx")
def export_xlsx_corpus(corpus_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        if not session.query(MembreCorpus).filter_by(corpus_id=corpus_id, utilisateur_id=user.id).first():
            raise HTTPException(403, "Tu n'as pas accès à ce corpus.")
        c = session.get(Corpus, corpus_id)
        if not c or not c.analyse:
            raise HTTPException(404, "Aucune analyse de corpus disponible à exporter.")
        entretiens = [_entretien_vers_dict(e) for e in session.query(Entretien).filter_by(corpus_id=corpus_id, statut="termine").order_by(Entretien.cree_le).all()]
        fiabilite = _fiabilite_corpus(session, corpus_id)
        buf = generer_xlsx_etude(_corpus_vers_dict(c), entretiens, fiabilite)
        nom = _nom_fichier(c.nom)
    finally:
        session.close()
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="rapport_{nom}.xlsx"'},
    )


@app.post("/api/transcriptions/{entretien_id}/analyser")
def lancer_analyse(
    entretien_id: str, contexte: str = Form(""), methode: str = Form("gioia"), langue: str = Form("fr"),
    user: Utilisateur = Depends(utilisateur_courant),
):
    if methode not in METHODES_ANALYSE:
        raise HTTPException(422, f"Méthode inconnue : {methode}")
    if langue not in ("fr", "en"):
        langue = "fr"
    session = get_session()
    try:
        e = session.get(Entretien, entretien_id)
        if not e:
            raise HTTPException(404, "Entretien introuvable.")
        _verifier_acces(session, e, user)
        if e.statut != "termine":
            raise HTTPException(409, "La transcription doit être terminée avant de lancer l'analyse.")
        if not e.segments:
            raise HTTPException(422, "Aucune parole détectée : rien à analyser.")
        if not user.est_admin and user.credits < COUT_CREDIT_ANALYSE:
            raise HTTPException(402, "Crédits insuffisants. Consulte les forfaits disponibles dans l'app.")
        if not user.est_admin:
            u = session.get(Utilisateur, user.id)
            u.credits -= COUT_CREDIT_ANALYSE
            session.add(MouvementCredit(
                id=uuid.uuid4().hex[:16], utilisateur_id=user.id, delta=-COUT_CREDIT_ANALYSE,
                motif=f"Analyse ({methode}) — {e.titre}",
            ))
        e.analyse_statut = "en_cours"
        e.analyse_erreur = None
        session.commit()
    finally:
        session.close()

    threading.Thread(target=_run_analyse, args=(entretien_id, contexte, methode, user.id, langue), daemon=True).start()
    return {"analyse_statut": "en_cours"}


# ----------------------------------------------------------------- codage collaboratif
@app.post("/api/transcriptions/{entretien_id}/codages")
def enregistrer_codage(entretien_id: str, payload: CodageIn, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        e = session.get(Entretien, entretien_id)
        if not e:
            raise HTTPException(404, "Entretien introuvable.")
        _verifier_acces(session, e, user)

        existant = session.query(Codage).filter_by(
            entretien_id=entretien_id, utilisateur_id=user.id, segment_index=payload.segment_index
        ).first()
        if existant:
            existant.code = payload.code
        else:
            session.add(Codage(
                id=uuid.uuid4().hex[:16], entretien_id=entretien_id, utilisateur_id=user.id,
                segment_index=payload.segment_index, code=payload.code,
            ))
        session.commit()
        return {"ok": True}
    finally:
        session.close()


@app.get("/api/transcriptions/{entretien_id}/codages")
def lister_codages(entretien_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    session = get_session()
    try:
        e = session.get(Entretien, entretien_id)
        if not e:
            raise HTTPException(404, "Entretien introuvable.")
        _verifier_acces(session, e, user)
        codages = session.query(Codage).filter_by(entretien_id=entretien_id).all()
        out = []
        for c in codages:
            u = session.get(Utilisateur, c.utilisateur_id)
            out.append({"segment_index": c.segment_index, "code": c.code, "codeur": u.nom or u.email if u else "?"})
        return out
    finally:
        session.close()


@app.get("/api/transcriptions/{entretien_id}/fiabilite")
def fiabilite_inter_codeurs(entretien_id: str, user: Utilisateur = Depends(utilisateur_courant)):
    """Calcule le kappa de Cohen entre chaque paire de codeurs ayant codé les mêmes segments."""
    session = get_session()
    try:
        e = session.get(Entretien, entretien_id)
        if not e:
            raise HTTPException(404, "Entretien introuvable.")
        _verifier_acces(session, e, user)
        return _fiabilite_entretien(session, entretien_id)
    finally:
        session.close()


def _fiabilite_entretien(session, entretien_id: str) -> dict:
    codages = session.query(Codage).filter_by(entretien_id=entretien_id).all()
    par_codeur: dict = {}
    for c in codages:
        par_codeur.setdefault(c.utilisateur_id, {})[c.segment_index] = c.code

    paires = []
    for u1, u2 in combinations(par_codeur.keys(), 2):
        communs = set(par_codeur[u1]) & set(par_codeur[u2])
        if len(communs) < 2:
            continue
        kappa = _kappa_cohen(
            [par_codeur[u1][i] for i in communs],
            [par_codeur[u2][i] for i in communs],
        )
        paires.append({"segments_communs": len(communs), "kappa": round(kappa, 3)})

    moyenne = round(sum(p["kappa"] for p in paires) / len(paires), 3) if paires else None
    return {"nb_codeurs": len(par_codeur), "paires": paires, "kappa_moyen": moyenne}


def _fiabilite_corpus(session, corpus_id: str) -> dict:
    """Agrège la fiabilité inter-codeurs sur tous les entretiens terminés d'un corpus,
    pour l'inclure dans le rapport d'étude complet."""
    entretiens = session.query(Entretien).filter_by(corpus_id=corpus_id, statut="termine").all()
    details = []
    for e in entretiens:
        f = _fiabilite_entretien(session, e.id)
        if f["nb_codeurs"] >= 2:
            details.append({"titre": e.titre, "nb_codeurs": f["nb_codeurs"], "kappa_moyen": f["kappa_moyen"]})
    return {"details": details}


def _kappa_cohen(a: list, b: list) -> float:
    n = len(a)
    accord_observe = sum(1 for x, y in zip(a, b) if x == y) / n
    labels = set(a) | set(b)
    accord_attendu = sum((a.count(lbl) / n) * (b.count(lbl) / n) for lbl in labels)
    if accord_attendu == 1:
        return 1.0
    return (accord_observe - accord_attendu) / (1 - accord_attendu)


# ----------------------------------------------------------------- moteur de transcription
def _run_job(entretien_id: str, path: str, langue: str, vocabulaire: str):
    session = get_session()
    try:
        e = session.get(Entretien, entretien_id)
        e.statut = "en_cours"
        session.commit()

        locuteurs = []
        try:
            pipeline = get_diarisation()
            diarisation = pipeline(path)
            for segment, _, label in diarisation.itertracks(yield_label=True):
                locuteurs.append({"debut": round(segment.start, 2), "fin": round(segment.end, 2), "locuteur": label})
        except Exception:  # noqa: BLE001
            pass  # diarisation facultative : on continue sans elle

        model = get_whisper()
        prompt = f"Entretien de recherche. Termes attendus : {vocabulaire}" if vocabulaire else None
        segments, info = model.transcribe(
            path,
            language=None if langue in ("auto", "dyu", "bci") else langue,
            initial_prompt=prompt,
            word_timestamps=True,
            vad_filter=True,
            # Réduction des hallucinations et gain d'exhaustivité, en particulier pour
            # les entretiens longs avec pauses, bruit de fond, et alternance de langues :
            # - condition_on_previous_text=False évite au modèle de "dériver" ou de
            #   boucler sur une phrase après un silence ou un changement de langue.
            # - hallucination_silence_threshold ignore les segments hallucinés générés
            #   sur du silence pur, sans jamais couper une parole réelle (vad_filter
            #   s'en charge séparément).
            # - beam_size=5 (recherche en faisceau) et température multi-passe
            #   (valeur par défaut de la bibliothèque) maximisent la fidélité au lieu
            #   d'une simple prédiction gloutonne.
            condition_on_previous_text=False,
            hallucination_silence_threshold=2.0,
            beam_size=5,
            compression_ratio_threshold=2.4,
            log_prob_threshold=-1.0,
            no_speech_threshold=0.6,
        )
        out = []
        for s in segments:
            out.append({
                "debut": round(s.start, 2),
                "fin": round(s.end, 2),
                "texte": s.text.strip(),
                "mots": [
                    {"mot": w.word.strip(), "confiance": round(w.probability, 2)}
                    for w in (s.words or [])
                ],
            })

        e = session.get(Entretien, entretien_id)
        e.segments = out
        e.locuteurs = locuteurs
        e.langue_detectee = info.language
        e.statut = "termine"
        if langue in ("dyu", "bci"):
            e.note = (
                "Langue locale demandée : la transcription utilise le modèle général. "
                "Un modèle affiné améliorerait sensiblement la qualité."
            )
        session.commit()
    except Exception as ex:  # noqa: BLE001
        e = session.get(Entretien, entretien_id)
        if e:
            e.statut = "erreur"
            e.erreur = str(ex)
            u = session.get(Utilisateur, e.proprietaire_id)
            if u and not u.est_admin:
                u.credits += COUT_CREDIT_TRANSCRIPTION
                session.add(MouvementCredit(
                    id=uuid.uuid4().hex[:16], utilisateur_id=u.id, delta=COUT_CREDIT_TRANSCRIPTION,
                    motif="Remboursement — échec de la transcription",
                ))
            session.commit()
    finally:
        session.close()
        try:
            os.unlink(path)
        except OSError:
            pass


# ----------------------------------------------------------------- analyse qualitative
SCHEMA_ANALYSE = """{
  "demarche_methodologique": "paragraphe de niveau doctoral (8-12 phrases) situant le positionnement épistémologique, justifiant le choix de la méthode pour ce corpus, et détaillant précisément la procédure suivie (étapes, critères de regroupement, itérations)",
  "premier_ordre": [{"concept": "libellé proche du verbatim", "verbatims": [{"texte": "citation exacte", "debut": 12.4, "entretien": "titre de l'entretien si plusieurs"}]}],
  "second_ordre": [{"theme": "libellé du regroupement de second niveau", "concepts_lies": ["concept 1", "concept 2"], "description": "1-2 phrases"}],
  "dimensions_agregees": [{"dimension": "libellé de la catégorie la plus abstraite", "themes_lies": ["thème 1", "thème 2"]}],
  "synthese": "synthèse interprétative de 4-6 phrases reliant les résultats à la question de recherche",
  "limites": "limites méthodologiques de ce codage automatique, en 2-3 phrases"
}"""

METHODES_ANALYSE = {
    "gioia": {
        "label": "Méthode Gioia (structuration des données)",
        "reference": "Gioia, Corley & Hamilton (2013)",
        "instructions": """Suis la méthode de structuration des données de Gioia, Corley & Hamilton (2013), \
le standard des revues de recherche en management (Academy of Management Journal, Organization Science) :
1. Premier ordre : concepts proches du langage des participants (in vivo, très proches du verbatim), \
chacun appuyé par un ou plusieurs verbatims exacts avec leur horodatage.
2. Second ordre : thèmes plus abstraits regroupant les concepts de premier ordre — c'est ici que le \
chercheur introduit son vocabulaire théorique.
3. Dimensions agrégées : catégories théoriques finales regroupant les thèmes de second ordre, formant \
la "data structure" typique d'un article Gioia.
Utilise TOUJOURS les trois niveaux (premier_ordre, second_ordre, dimensions_agregees).""",
    },
    "thematique": {
        "label": "Analyse thématique réflexive (Braun & Clarke)",
        "reference": "Braun & Clarke (2006, 2019)",
        "instructions": """Suis les six phases de l'analyse thématique réflexive de Braun & Clarke (2006, 2019) : \
familiarisation avec les données, génération de codes initiaux, recherche de thèmes, révision des thèmes, \
définition et nommage des thèmes, production du rapport.
Mets les codes initiaux dans "premier_ordre" (chaque code appuyé par ses verbatims), et les thèmes \
définitifs (regroupements de codes, avec leur description) dans "second_ordre". Cette méthode ne \
comporte PAS de troisième niveau d'abstraction théorique : laisse "dimensions_agregees" comme une \
liste VIDE []. Insiste dans "demarche_methodologique" sur la nature réflexive et itérative du \
processus (le chercheur comme instrument actif de l'analyse), propre à cette méthode.""",
    },
    "contenu": {
        "label": "Analyse de contenu catégorielle (Bardin)",
        "reference": "Bardin (1977/2013)",
        "instructions": """Suis l'analyse de contenu catégorielle de Bardin : pré-analyse, exploitation du \
matériel par découpage en unités de sens, catégorisation, et inférence.
Mets les unités de sens (unités d'enregistrement) dans "premier_ordre", les sous-catégories qui les \
regroupent dans "second_ordre", et les catégories finales dans "dimensions_agregees". Pour chaque \
sous-catégorie, indique dans sa "description" sa fréquence d'apparition dans le corpus (nombre \
d'occurrences) — c'est une dimension centrale de cette méthode, plus quantitative que les deux autres. \
Précise dans "demarche_methodologique" les règles de découpage retenues (thématique, par énoncé...) \
et les critères de catégorisation (exclusion mutuelle, homogénéité, pertinence).""",
    },
}


CONSIGNE_ORIGINALITE = (
    "Consigne d'originalité (essentielle — cette application est utilisée par des étudiants et "
    "chercheurs dont les productions peuvent être comparées entre elles ou passées dans un logiciel "
    "anti-plagiat) : rédige un contenu ORIGINAL, formulé avec tes propres mots, et directement ancré "
    "dans les éléments concrets et spécifiques de CE cas précis (ce thème, ce contexte, ces verbatims). "
    "N'utilise jamais de tournure figée, de définition récitée mot pour mot, ou de paragraphe "
    "interchangeable qui pourrait convenir à l'identique à un autre thème ou un autre chercheur — même "
    "les passages méthodologiques génériques (démarche, conseils) doivent être reformulés à ta façon à "
    "chaque fois plutôt que recopiés d'un modèle standard. Les citations verbatim exactes du terrain "
    "restent seules exemptées de cette règle : elles doivent être retranscrites fidèlement, jamais "
    "reformulées."
)


def _construire_prompt(transcription: str, contexte: str, methode: str, multi_entretiens: bool = False, langue: str = "fr") -> str:
    m = METHODES_ANALYSE[methode]
    est_anglais = langue == "en"
    portee = (
        "sur l'ensemble des entretiens de ce corpus (analyse transversale inter-cas — indique pour "
        "chaque verbatim de quel entretien il provient via le champ \"entretien\", et commente dans "
        "la synthèse le degré de convergence/divergence entre entretiens ainsi que si une saturation "
        "théorique semble atteinte ou si de nouveaux codes émergent encore)"
        if multi_entretiens else
        "de la transcription d'entretien ci-dessous"
    )
    consigne_langue = (
        "Rédige tous les libellés (concepts, thèmes, dimensions), la démarche méthodologique, la "
        "synthèse et les limites en ANGLAIS. En revanche, les verbatims (citations) doivent "
        "impérativement rester dans leur langue EXACTE d'origine telle qu'elle apparaît dans la "
        "transcription (français, dioula, baoulé, anglais...) — ne traduis JAMAIS un verbatim, "
        "traduire les paroles d'un participant fausserait la rigueur scientifique de l'analyse."
        if est_anglais else
        "Rédige l'intégralité de la réponse en français, y compris les verbatims tels qu'ils "
        "apparaissent dans la transcription (ne les traduis pas s'ils sont dans une autre langue)."
    )
    return f"""Tu es méthodologue qualitatif spécialisé en sciences de gestion et organisation, de niveau \
recherche doctorale.

Effectue un codage {portee}.

{m['instructions']}

Contexte / question de recherche fournie par le chercheur : {contexte or "non précisée"}

Transcription{'s' if multi_entretiens else ''} horodatée{'s' if multi_entretiens else ''} :
{transcription}

{consigne_langue}

{CONSIGNE_ORIGINALITE}

Réponds UNIQUEMENT avec un objet JSON strictement conforme à ce schéma — aucun texte avant ou après, \
aucune balise markdown. Assure-toi que le JSON est syntaxiquement valide : échappe tous \
les guillemets internes aux chaînes de caractères (\\") et ne laisse aucune chaîne non terminée :
{SCHEMA_ANALYSE}"""


def _echapper_controles_dans_chaines(texte: str) -> str:
    """Corrige les sauts de ligne/tabulations LITTÉRAUX insérés par erreur à
    l'intérieur des valeurs de chaîne JSON — le modèle doit les échapper en \\n,
    mais l'omet parfois sur un texte long en plusieurs paragraphes. Sans ce
    correctif, ces caractères de contrôle cassent le parsing JSON strict, et
    l'ancien mécanisme de réparation (pensé pour une vraie troncature) tronquait
    alors le contenu bien plus tôt que nécessaire, perdant silencieusement des
    champs pourtant bien présents dans la réponse du modèle."""
    resultat = []
    dans_chaine = False
    echappement = False
    for ch in texte:
        if echappement:
            resultat.append(ch)
            echappement = False
            continue
        if ch == "\\":
            resultat.append(ch)
            echappement = True
            continue
        if ch == '"':
            dans_chaine = not dans_chaine
            resultat.append(ch)
            continue
        if dans_chaine and ch == "\n":
            resultat.append("\\n")
            continue
        if dans_chaine and ch == "\t":
            resultat.append("\\t")
            continue
        if dans_chaine and ch == "\r":
            continue
        resultat.append(ch)
    return "".join(resultat)


def _extraire_json(texte: str) -> dict:
    """Nettoie et parse la réponse du modèle, avec une tentative de réparation si le JSON
    a été coupé net (troncature liée à la limite de tokens)."""
    texte = texte.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
    texte = _echapper_controles_dans_chaines(texte)
    try:
        return json.loads(texte)
    except json.JSONDecodeError:
        # Réparation simple : si la chaîne est coupée en plein milieu, on referme au dernier
        # point de structure valide (dernière accolade/crochet fermé équilibré).
        for fin in range(len(texte), 0, -1):
            fragment = texte[:fin]
            if fragment.count("{") == 0:
                break
            try:
                return json.loads(fragment + "}" * (fragment.count("{") - fragment.count("}")))
            except json.JSONDecodeError:
                continue
        raise


def _valider_analyse(analyse: dict):
    for cle in ("premier_ordre", "second_ordre", "dimensions_agregees", "synthese", "limites"):
        if cle not in analyse:
            raise ValueError(f"Réponse du modèle incomplète (« {cle} » manquant).")


def _run_analyse(entretien_id: str, contexte: str, methode: str, payeur_id: str, langue: str = "fr"):
    session = get_session()
    try:
        e = session.get(Entretien, entretien_id)
        segments = list(e.segments or [])
        transcription = "\n".join(f"[{s['debut']:.1f}s] {s['texte']}" for s in segments)
        client = get_anthropic()
        prompt = _construire_prompt(transcription, contexte, methode, langue=langue)
        resp = client.messages.create(
            model=ANALYSE_MODEL,
            max_tokens=8000,
            messages=[{"role": "user", "content": prompt}],
        )
        texte = "".join(b.text for b in resp.content if getattr(b, "type", None) == "text")
        analyse = _extraire_json(texte)
        _valider_analyse(analyse)

        e.analyse = analyse
        e.analyse_statut = "termine"
        e.analyse_contexte = contexte
        e.analyse_methode = methode
        e.analyse_modele = ANALYSE_MODEL
        e.analyse_langue = langue
        session.commit()
    except Exception as ex:  # noqa: BLE001
        e = session.get(Entretien, entretien_id)
        if e:
            e.analyse_statut = "erreur"
            e.analyse_erreur = str(ex)
            u = session.get(Utilisateur, payeur_id)
            if u and not u.est_admin:
                u.credits += COUT_CREDIT_ANALYSE
                session.add(MouvementCredit(
                    id=uuid.uuid4().hex[:16], utilisateur_id=u.id, delta=COUT_CREDIT_ANALYSE,
                    motif="Remboursement — échec de l'analyse",
                ))
            session.commit()
    finally:
        session.close()


def _run_analyse_corpus(corpus_id: str, contexte: str, methode: str, payeur_id: str, cout: float, langue: str = "fr"):
    session = get_session()
    try:
        entretiens = session.query(Entretien).filter_by(corpus_id=corpus_id, statut="termine").all()
        blocs = []
        for e in entretiens:
            for s in (e.segments or []):
                blocs.append(f"[{e.titre} · {s['debut']:.1f}s] {s['texte']}")
        transcription = "\n".join(blocs)

        client = get_anthropic()
        prompt = _construire_prompt(transcription, contexte, methode, multi_entretiens=True, langue=langue)
        resp = client.messages.create(
            model=ANALYSE_MODEL,
            max_tokens=8000,
            messages=[{"role": "user", "content": prompt}],
        )
        texte = "".join(b.text for b in resp.content if getattr(b, "type", None) == "text")
        analyse = _extraire_json(texte)
        _valider_analyse(analyse)

        corpus = session.get(Corpus, corpus_id)
        corpus.analyse = analyse
        corpus.analyse_statut = "termine"
        corpus.analyse_contexte = contexte
        corpus.analyse_methode = methode
        corpus.analyse_modele = ANALYSE_MODEL
        corpus.analyse_nb_entretiens = len(entretiens)
        corpus.analyse_langue = langue
        session.commit()
    except Exception as ex:  # noqa: BLE001
        corpus = session.get(Corpus, corpus_id)
        if corpus:
            corpus.analyse_statut = "erreur"
            corpus.analyse_erreur = str(ex)
            u = session.get(Utilisateur, payeur_id)
            if u and not u.est_admin:
                u.credits += cout
                session.add(MouvementCredit(
                    id=uuid.uuid4().hex[:16], utilisateur_id=u.id, delta=cout,
                    motif="Remboursement — échec de l'analyse transversale",
                ))
            session.commit()
    finally:
        session.close()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", "8000")))
